"""
BOM Generation Engine v3.0
──────────────────────────
4 core capabilities:
  1. Structured BOM with component-level material specifications
  2. Sub-assembly grouping from real dissection data
  3. Material traceability — MOC linked to fluid-temperature rules
  4. Weight schedule per sub-assembly for foundation & crane selection

Author : Ayush Kamle
Stack  : Pure Python — pandas, openpyxl, re, math, json
"""

import os, re, math, json, datetime
import pandas as pd
from io import BytesIO

# ─────────────────────────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────────────────────────
_DIR     = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(_DIR, "Component_Library_COMPLETE.xlsx")
LRN_PATH = os.path.join(_DIR, "learning_data.json")

# ─────────────────────────────────────────────────────────────────
# SUB-ASSEMBLY HIERARCHY MAP
# Every category maps to: (section, level, assembly_group)
# Level 1 = package, 2 = sub-assembly, 3 = component
# ─────────────────────────────────────────────────────────────────
HIERARCHY = {
    # Section A — Pump Hydraulics
    "Pump":          ("A. PUMP HYDRAULICS",        1, "Pump Assembly"),
    "Casing":        ("A. PUMP HYDRAULICS",        2, "Casing Assembly"),
    "Impeller":      ("A. PUMP HYDRAULICS",        3, "Casing Assembly"),
    "Liner":         ("A. PUMP HYDRAULICS",        3, "Casing Assembly"),
    "Wear Ring":     ("A. PUMP HYDRAULICS",        3, "Casing Assembly"),
    "Rotor":         ("A. PUMP HYDRAULICS",        2, "Rotor Assembly"),
    # Section B — Rotating Assembly
    "Shaft":         ("B. ROTATING ASSEMBLY",      3, "Rotor Assembly"),
    "Sleeve":        ("B. ROTATING ASSEMBLY",      3, "Rotor Assembly"),
    # Section C — Bearings
    "Bearing":       ("C. BEARINGS & LUBRICATION", 3, "Bearing Assembly"),
    "Housing":       ("C. BEARINGS & LUBRICATION", 2, "Bearing Assembly"),
    "Lubrication":   ("C. BEARINGS & LUBRICATION", 3, "Bearing Assembly"),
    "Oiler":         ("C. BEARINGS & LUBRICATION", 3, "Bearing Assembly"),
    # Section D — Sealing
    "Seal":          ("D. SHAFT SEALING",          2, "Sealing Assembly"),
    "Mechanical Seal":("D. SHAFT SEALING",         2, "Sealing Assembly"),
    "Gland":         ("D. SHAFT SEALING",          3, "Sealing Assembly"),
    "Stuffing Box":  ("D. SHAFT SEALING",          3, "Sealing Assembly"),
    "Shim":          ("D. SHAFT SEALING",          3, "Sealing Assembly"),
    # Section E — Drive
    "Coupling":      ("E. DRIVE & COUPLING",       2, "Drive Assembly"),
    "Guard":         ("E. DRIVE & COUPLING",       3, "Drive Assembly"),
    "V-Belt":        ("E. DRIVE & COUPLING",       3, "Drive Assembly"),
    "Pulley":        ("E. DRIVE & COUPLING",       3, "Drive Assembly"),
    "Belt":          ("E. DRIVE & COUPLING",       3, "Drive Assembly"),
    # Section F — Motor
    "Motor":         ("F. MOTOR / DRIVER",         1, "Motor"),
    # Section G — Structural
    "Baseplate":     ("G. STRUCTURAL",             1, "Structural"),
    "Foundation":    ("G. STRUCTURAL",             2, "Structural"),
    "Stool":         ("G. STRUCTURAL",             2, "Structural"),
    "Saddle":        ("G. STRUCTURAL",             2, "Structural"),
    "Frame":         ("G. STRUCTURAL",             2, "Structural"),
    "Bracket":       ("G. STRUCTURAL",             2, "Structural"),
    # Section H — Piping & Nozzles
    "Flange":        ("H. PIPING & NOZZLES",       2, "Piping Assembly"),
    "Piping":        ("H. PIPING & NOZZLES",       3, "Piping Assembly"),
    "Column":        ("H. PIPING & NOZZLES",       2, "Piping Assembly"),
    "Strainer":      ("H. PIPING & NOZZLES",       3, "Piping Assembly"),
    "Nozzle":        ("H. PIPING & NOZZLES",       2, "Piping Assembly"),
    # Section I — Fasteners & Gaskets
    "Fastener":      ("I. FASTENERS & GASKETS",    3, "Fasteners"),
    "Fasteners":     ("I. FASTENERS & GASKETS",    3, "Fasteners"),
    "Gasket":        ("I. FASTENERS & GASKETS",    3, "Fasteners"),
    "Bolt":          ("I. FASTENERS & GASKETS",    3, "Fasteners"),
    # Section J — Instrumentation
    "Instrumentation":("J. INSTRUMENTATION",       2, "Instrumentation"),
    "Thermometer":   ("J. INSTRUMENTATION",        3, "Instrumentation"),
    "Gauge":         ("J. INSTRUMENTATION",        3, "Instrumentation"),
    # Section K — Acoustic & Safety
    "Enclosure":     ("K. ACOUSTIC & SAFETY",      1, "Acoustic Enclosure"),
    "Acoustic":      ("K. ACOUSTIC & SAFETY",      1, "Acoustic Enclosure"),
    # Section L — Complete Assembly
    "Assembly":      ("L. COMPLETE ASSEMBLY",      1, "Complete Package"),
}

SECTION_ORDER = [
    "A. PUMP HYDRAULICS",
    "B. ROTATING ASSEMBLY",
    "C. BEARINGS & LUBRICATION",
    "D. SHAFT SEALING",
    "E. DRIVE & COUPLING",
    "F. MOTOR / DRIVER",
    "G. STRUCTURAL",
    "H. PIPING & NOZZLES",
    "I. FASTENERS & GASKETS",
    "J. INSTRUMENTATION",
    "K. ACOUSTIC & SAFETY",
    "L. COMPLETE ASSEMBLY",
]

# ─────────────────────────────────────────────────────────────────
# WEIGHT SCHEDULE DATA — from real dissection sheets
# ─────────────────────────────────────────────────────────────────
REAL_WEIGHTS = {
    "FLW-001": {
        "Pump (bare)":              1160,
        "Rotor Assembly":            173,
        "Top Casing Half":           272,
        "Motor (CGL 550kW)":        5300,
        "Baseplate (IS2062)":       3250,
        "Coupling (Rathi)":           50,
        "Accessories":                50,
        "Acoustic Enclosure (ARK)": 2350,
        "TOTAL PACKAGE":           12160,
        "Heaviest Single Lift":      272,
        "Heaviest Lift Item":    "Top Casing Half",
    },
    "METSO-001": {
        "Pump (bare)":              1010,
        "Motor (Innomotics 500kW)": 3400,
        "V-Belt Guard":              107,
        "Bed Frame with Motor":      792,
        "TOTAL PACKAGE":            5390,
        "Heaviest Single Lift":     1010,
        "Heaviest Lift Item":   "Pump bare",
    },
    "WILO-001": {
        "Pump Assembly":            2175,
        "Motor (KECL 200kW)":       1350,
        "TOTAL PACKAGE":            2175,
        "Heaviest Single Lift":     2175,
        "Heaviest Lift Item":   "Complete pump assembly",
    },
    "WILO-002": {
        "Pump Assembly":            1823,
        "Motor (190kW)":            1250,
        "TOTAL PACKAGE":            1823,
        "Heaviest Single Lift":     1823,
        "Heaviest Lift Item":   "Complete pump assembly",
    },
    "JYOTI-001": {
        "Pump (VTP complete)":      5280,
        "Motor (315kW)":            2800,
        "TOTAL PACKAGE":            8080,
        "Heaviest Single Lift":     5280,
        "Heaviest Lift Item":   "VTP pump assembly",
    },
    "JYOTI-002": {
        "Pump (VTP complete)":      3350,
        "Motor (160kW)":            1350,
        "TOTAL PACKAGE":            4700,
        "Heaviest Single Lift":     3350,
        "Heaviest Lift Item":   "VTP pump assembly",
    },
    "KSB-001": {
        "Pump + Column Assembly":    410,
        "Motor (45kW)":              750,
        "TOTAL PACKAGE":             410,
        "Heaviest Single Lift":      410,
        "Heaviest Lift Item":   "Pump assembly",
    },
    "KSB-003": {
        "Pump barrel + column":     1390,
        "Motor (ABB 30kW)":          171,
        "Motor lantern + base":      170,
        "TOTAL PACKAGE":            1561,
        "Heaviest Single Lift":     1390,
        "Heaviest Lift Item":   "Pump barrel + column pipe",
    },
    "KSB-004": {
        "Pump (empty)":              193,
        "Motor (ABB 15kW)":          126,
        "TOTAL PACKAGE":             319,
        "Heaviest Single Lift":      193,
        "Heaviest Lift Item":   "Pump assembly (empty)",
    },
    # From Pump_Master_List (database record weights)
    "KSB-002": {
        "Pump assembly":             350,
        "Motor (5.5kW)":              48,
        "TOTAL PACKAGE":             398,
        "Heaviest Single Lift":      350,
        "Heaviest Lift Item":   "Pump assembly",
    },
    "KSB-005": {
        "Pump (VS6 double-casing)":  420,
        "Motor (VTA)":               150,
        "TOTAL PACKAGE":             570,
        "Heaviest Single Lift":      420,
        "Heaviest Lift Item":   "VS6 pump assembly (estimated)",
        "_note":                     "Weight VTA — estimate based on similar VS6",
    },
    "WILO-003": {
        "Pump (MPS-3 VTP 1-stage)":  680,
        "Motor (CGL VPC710 260kW)":  820,
        "Base frame + motor stool":  240,
        "TOTAL PACKAGE":            1740,
        "Heaviest Single Lift":      820,
        "Heaviest Lift Item":   "Motor (CGL VPC710)",
        "_note":                     "Estimate — GA weight not in dissection sheet",
    },
}

# Crane selection rules
def crane_category(total_kg):
    if total_kg <= 500:   return "Chain block / fork lift"
    if total_kg <= 3000:  return "Mobile crane 5–10T required"
    if total_kg <= 10000: return "Mobile crane 25T required"
    if total_kg <= 30000: return "Mobile crane 50T required"
    return "Heavy lift crane — route survey required"

def foundation_loads(total_kg):
    static  = total_kg
    dynamic = round(total_kg * 0.15)
    return {"static_kg": static, "dynamic_kg": dynamic,
            "note": "Dynamic = ±15% of static (general rule)"}


# ═══════════════════════════════════════════════════════════════════
# SECTION 1 — DATABASE LOADER
# ═══════════════════════════════════════════════════════════════════

def load_db():
    return {
        "pumps":      pd.read_excel(DB_PATH, sheet_name="Pump_Master_List"),
        "comps":      pd.read_excel(DB_PATH, sheet_name="Component_Library"),
        "mats":       pd.read_excel(DB_PATH, sheet_name="Material_Database"),
        "vendors":    pd.read_excel(DB_PATH, sheet_name="Vendor_Database"),
        "bom_tpl":    pd.read_excel(DB_PATH, sheet_name="BOM_Templates",          header=4),
        "physics":    pd.read_excel(DB_PATH, sheet_name="Physics_Parameters",     header=4),
        "mat_compat": pd.read_excel(DB_PATH, sheet_name="Material_Compatibility", header=4),
    }


# ═══════════════════════════════════════════════════════════════════
# SECTION 2 — LEARNING STORE
# ═══════════════════════════════════════════════════════════════════

_DEFAULT = {
    "feedback": [], "patterns": [], "corrections": [],
    "weight_calibs": {},
    "stats": {"total_sessions":0,"tier1_hits":0,"tier2_hits":0,
              "corrections":0,"patterns_added":0},
}

def get_store():
    if os.path.exists(LRN_PATH):
        try:
            with open(LRN_PATH) as f:
                data = json.load(f)
            for k, v in _DEFAULT.items():
                if k not in data: data[k] = v
            return data
        except Exception:
            pass
    return {k: (v.copy() if isinstance(v,dict) else list(v))
            for k,v in _DEFAULT.items()}

def _save(store):
    try:
        with open(LRN_PATH,"w") as f:
            json.dump(store, f, indent=2, default=str)
    except Exception:
        pass

def log_feedback(specs, bom_df, tier, confirmed_pump_type,
                 confirmed_moc, confirmed_weights, notes=""):
    store = get_store()
    store["feedback"].append({
        "ts": datetime.datetime.now().isoformat(),
        "specs": {k:v for k,v in specs.items() if v is not None},
        "tier": tier, "pump_type": confirmed_pump_type,
        "moc": confirmed_moc, "weights": confirmed_weights,
        "bom_rows": len(bom_df), "notes": notes,
        "ns": _ns(specs),
    })
    store["stats"]["total_sessions"] += 1
    if tier == "tier1": store["stats"]["tier1_hits"] += 1
    else:               store["stats"]["tier2_hits"]  += 1
    _update_calibs(store, confirmed_pump_type, confirmed_weights, specs)
    _save(store)

def log_pattern(field, bad, correct, snippet, notes=""):
    store = get_store()
    store["patterns"].append({
        "ts": datetime.datetime.now().isoformat(),
        "field": field, "bad": str(bad), "correct": str(correct),
        "snippet": snippet[:200], "notes": notes,
    })
    store["stats"]["patterns_added"] += 1
    _save(store)

def log_correction(specs, wrong, correct, notes=""):
    store = get_store()
    store["corrections"].append({
        "ts": datetime.datetime.now().isoformat(),
        "ns": _ns(specs), "fluid": (specs.get("fluid") or ""),
        "wrong_type": wrong, "correct_type": correct, "notes": notes,
    })
    store["stats"]["corrections"] += 1
    _save(store)

def get_learned_correction(Ns, fluid):
    store = get_store()
    fl = (fluid or "").lower()
    for c in reversed(store.get("corrections",[])):
        c_ns = c.get("ns"); c_fl = (c.get("fluid") or "").lower()
        ns_ok = not c_ns or not Ns or abs(c_ns-Ns)/max(Ns,1) < 0.30
        fl_ok = any(w in c_fl for w in fl.split()[:3]) or \
                any(w in fl   for w in c_fl.split()[:3])
        if ns_ok and fl_ok and c.get("correct_type"):
            return c["correct_type"]
    return None

def _ns(specs):
    Q=specs.get("flow_m3h"); H=specs.get("head_m"); n=specs.get("speed_rpm",1450)
    if Q and H and Q>0 and H>0:
        try: return round(calc_specific_speed(Q,H,n),1)
        except: pass
    return None

def _update_calibs(store, pump_type, weights, specs):
    if not weights: return
    pt  = (pump_type or "").lower()[:20]
    kw  = float(specs.get("motor_kw") or 30)
    if pt not in store["weight_calibs"]:
        store["weight_calibs"][pt] = {"pump_coeff":1.0,"motor_coeff":1.0,"n_samples":0}
    cal = store["weight_calibs"][pt]; n = cal["n_samples"]+1
    pp = _base_pump_wt(pt,kw); pm = _base_motor_wt(kw)
    ap = float(weights.get("pump_kg",0) or 0)
    am = float(weights.get("motor_kg",0) or 0)
    if pp>0 and ap>0: cal["pump_coeff"]  = (cal["pump_coeff"]*(n-1)+(ap/pp))/n
    if pm>0 and am>0: cal["motor_coeff"] = (cal["motor_coeff"]*(n-1)+(am/pm))/n
    cal["n_samples"] = n
    _save(store)

def _base_pump_wt(pt,kw):
    if "slurry"  in pt: return 1.8*kw**0.85
    if "turbine" in pt: return 0.45*kw**0.9*6**0.3
    if "sump"    in pt: return 0.9*kw**0.85
    return 2.1*kw**0.72

def _base_motor_wt(kw):
    return 6.2*kw**0.80 if kw>200 else 8.5*kw**0.75


# ═══════════════════════════════════════════════════════════════════
# SECTION 3 — PDF EXTRACTION
# ═══════════════════════════════════════════════════════════════════

def extract_pdf_text(file_bytes):
    try:
        import pdfplumber
        pages = []
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            for p in pdf.pages:
                t = p.extract_text()
                if t: pages.append(t)
        return "\n".join(pages), None
    except Exception as e:
        return "", str(e)


# ═══════════════════════════════════════════════════════════════════
# SECTION 4 — SPEC PARSER
# ═══════════════════════════════════════════════════════════════════

_PATTERNS = {
    "flow_m3h": [
        r"(?:[Rr]ated\s+)?(?:[Vv]olumetric\s+)?[Ff]low\s*(?:[Rr]ate|[Cc]apacity)?\s*[:\-=]?\s*(\d+\.?\d*)\s*m3/h",
        r"[Ff]low\s*[:\-=]?\s*(\d+\.?\d*)\s*m\xb3/h",
        r"[Cc]apacity\s*[:\-=]?\s*(\d+\.?\d*)\s*m3/h",
        r"[Dd]ischarge\s*[:\-=]?\s*(\d+\.?\d*)\s*m3/h",
        r"\bQ\s*[=:\-]\s*(\d+\.?\d*)\s*m3",
        r"(\d+\.?\d*)\s*m3/hr",
        r"[Cc]apacity\s+of\s+each\s+pump\s+(?:m3/h[r]?)?\s*(\d+\.?\d*)",
        r"(?:Flow|Discharge|Capacity)\s+(?:m3/h[r]?|m\xb3/h[r]?)\s+(\d+\.?\d*)",
        r"\b(\d{3,5}\.?\d*)\s*LPM\b",
        r"\b(\d{3,5}\.?\d*)\s*L/[Mm]in\b",
        r"\b(\d+\.?\d*)\s*[Ll][Pp][Ss]\b",
        r"\b(\d+\.?\d*)\s*[Ll]/[Ss]\b",
    ],
    "head_m": [
        r"[Pp]ump\s+[Rr]ated\s+[Hh]ead\s*[:\-=]?\s*(\d+\.?\d*)\s*m[\b\s]",
        r"[Tt]otal\s+(?:[Dd]ynamic\s+)?[Hh]ead\s*[:\-=]?\s*(\d+\.?\d*)\s*m[\b\s]",
        r"[Rr]ated\s+[Hh]ead\s*[:\-=]?\s*(\d+\.?\d*)\s*m[\b\s]",
        r"\bTDH\s*[:\-=]?\s*(\d+\.?\d*)\s*m",
        r"\bH\s*[=:\-]\s*(\d+\.?\d*)\s*m\b",
        r"[Hh]ead\s*[:\-=]?\s*(\d+\.?\d*)\s*[Mm][Ww][Cc]",
        r"(\d+\.?\d*)\s*[Mm][Ww][Cc]",
        r"[Hh]ead\s*[:\-=]?\s*(\d+\.?\d*)\s*[Mm][Ll][Cc]",
        r"(\d+\.?\d*)\s*[Mm][Ll][Cc]",
        r"[Pp]ump\s+[Dd]ifferential\s+[Hh]ead[^\d]+(\d+\.?\d*)",
        r"[Dd]ifferential\s+[Hh]ead[^\d]+(\d+\.?\d*)",
        r"(\d{1,3}\.?\d*)\s*[Mm]tr\b",
        r"[Hh]ead\s*[:\-=]?\s*(\d+\.?\d*)\s*m\b",
    ],
    "speed_rpm": [
        r"[Ff]ull\s+[Ll]oad\s+[Ss]peed.*?(\d{3,4})\s*[Rr][Pp][Mm]",
        r"[Rr]ated\s+[Ss]peed.*?(\d{3,4})\s*[Rr][Pp][Mm]",
        r"[Ss]peed\s+of\s+[Pp]ump[^\d]+(\d{3,4})",
        r"[Ss]peed\s*[:\-=]?\s*(\d{3,4})\s*[Rr][Pp][Mm]",
        r"(\d{3,4})\s*[Rr][Pp][Mm]",
    ],
    "motor_kw": [
        r"[Mm]otor\s+[Rr]ating\s*[:\-=]?\s*(\d+\.?\d*)\s*[Kk][Ww]",
        r"[Mm]otor\s+[Pp]ower\s*[:\-=]?\s*(\d+\.?\d*)\s*[Kk][Ww]",
        r"[Nn]ominal\s+[Pp]ower\s*[:\-=]?\s*(\d+\.?\d*)\s*[Kk][Ww]",
        r"[Rr]ated\s+[Oo]utput.*?(\d+\.?\d*)\s*[Kk][Ww]",
        r"[Ss]elected\s+[Dd]rive\s+[Rr]ating.*?(\d+\.?\d*)\s*[Kk][Ww]",
        r"[Ss]elected\s+[Mm]otor.*?(\d+\.?\d*)\s*[Kk][Ww]",
        r"[Mm]otor\s*[:\-=]?\s*(\d+\.?\d*)\s*[Kk][Ww]",
        r"(\d+\.?\d*)\s*[Kk][Ww]\s*(?:motor|Motor|MOTOR|nameplate)",
        r"(\d+\.?\d*)\s*[Hh][Pp]\b",
    ],
    "temp_c": [
        r"[Oo]p(?:erating)?\s+[Tt]emp(?:erature)?\s*[:\-=]?\s*(\d+\.?\d*)\s*[°\xb0]?[Cc]",
        r"[Pp]rocess\s+[Tt]emp(?:erature)?\s*[:\-=]?\s*(\d+\.?\d*)\s*[°\xb0]?[Cc]",
        r"[Tt]emp(?:erature)?\s+[Oo]f\s+[Ff]luid[^\d]*(\d+\.?\d*)",
        r"[Tt]emp(?:erature)?\s*[:\-=]?\s*(\d+\.?\d*)\s*[°\xb0][Cc]",
        r"(\d{2,3})\s*[Dd]eg\.?\s*[Cc]\b",
        r"(\d{2,3})\s*[°\xb0][Cc]\b",
        r"[Ff]luid\s+[Tt]emp[^\d]+(\d+\.?\d*)",
    ],
    "density_kgm3": [
        r"[Dd]ensity\s*[:\-=]?\s*(\d{3,4}\.?\d*)\s*kg",
        r"[Ss]pecific\s+[Gg]ravity\s*[:\-=]?\s*(\d\.?\d+)",
        r"\bSG\s*[=:\-]\s*(\d\.?\d+)\b",
        r"\bS\.?G\.?\s*[=:\-]\s*(\d\.?\d+)\b",
        r"[Ss]p\.?\s*[Gg]r\.?\s*[=:\-]\s*(\d\.?\d+)",
    ],
    "stages": [
        r"[Nn]o\.?\s+[Oo]f\s+[Ss]tages?\s*[:\-=]?\s*(\d+)",
        r"[Nn]umber\s+[Oo]f\s+[Ss]tages?\s*[:\-=]?\s*(\d+)",
        r"(\d+)\s*[Ss]tage\s+[Pp]ump",
        r"(\d+)[- ][Ss]tage\b",
    ],
}

_FLUID_MAP = [
    ("live steam condensate",  "Live Steam Condensate"),
    ("steam condensate",       "Live Steam Condensate"),
    ("process condensate",     "Process Condensate"),
    ("condensate",             "Process Condensate"),
    ("caustic liquor",         "Caustic Liquor (Alumina)"),
    ("alumina liquor",         "Caustic Liquor (Alumina)"),
    ("caustic soda",           "Caustic Soda"),
    ("caustic",                "Caustic Liquor"),
    ("sulphuric acid",         "Dilute Sulphuric Acid"),
    ("acid",                   "Dilute Acid"),
    ("slurry",                 "Slurry"),
    ("seawater",               "Seawater"),
    ("sea water",              "Seawater"),
    ("crude oil",              "Crude Oil"),
    ("boiler feed",            "Boiler Feed Water"),
    ("cooling water",          "Cooling Water"),
    ("clear water",            "Clear Water"),
    ("clean water",            "Clear Water"),
    ("raw water",              "Clear Water"),
    ("water",                  "Clear Water"),
]

_DENS_DEF = {
    "Slurry":1300,"Caustic Liquor (Alumina)":1244,
    "Live Steam Condensate":930,"Process Condensate":990,
    "Dilute Sulphuric Acid":1050,"Seawater":1025,
    "Boiler Feed Water":950,"Crude Oil":870,"Cooling Water":998,
}

_SANITY = {
    "flow_m3h":(0,100000),"head_m":(0.5,2000),"speed_rpm":(200,10000),
    "motor_kw":(0.1,5000),"temp_c":(-10,600),"density_kgm3":(500,5000),
    "stages":(1,30),
}

def detect_multi_pump(text):
    """
    Detect if a PDF contains multiple distinct pump specifications.
    Returns list of {"label": str, "text": str} dicts — one per pump.
    Returns empty list if only one pump found (normal flow continues).

    Detection strategy:
      - Numbered pump sections: "1. Hydrant Pump", "2. Spray Pump"
      - Service headers: "SERVICE: Jockey Pump"
      - Repeated flow/head blocks — N occurrences of flow value = N pumps
    """
    t = text.replace("\r", "")

    # Strategy 1: numbered pump sections (most reliable)
    # Matches "1. Hydrant Water Pump", "2) Spray Pump", etc.
    numbered = re.findall(
        r"(?:^|\n)\s*(\d+[\.)\s]+[A-Z][A-Za-z ]{4,45}Pump)",
        t, re.MULTILINE
    )
    if len(numbered) >= 2:
        # Use their positions as segment boundaries
        candidates = []
        for m in re.finditer(
            r"(?:^|\n)\s*(\d+[\.)\s]+[A-Z][A-Za-z ]{4,45}Pump)",
            t, re.MULTILINE
        ):
            label = m.group(1).strip()
            candidates.append((m.start(), label))
        if len(candidates) >= 2:
            segments = []
            for i, (pos, label) in enumerate(candidates):
                end = candidates[i+1][0] if i+1 < len(candidates) else len(t)
                segments.append({"label": label, "text": t[pos:end]})
            return segments

    # Strategy 2: SERVICE / TAG labels repeating
    service_matches = re.findall(
        r"(?:Service|SERVICE|Tag|TAG|Equipment|EQUIPMENT)\s*[:\-]\s*([A-Za-z][A-Za-z ]{4,50})",
        t
    )
    if len(service_matches) >= 2:
        candidates = []
        for m in re.finditer(
            r"(?:Service|SERVICE|Tag|TAG)\s*[:\-]\s*([A-Za-z][A-Za-z ]{4,50})",
            t
        ):
            label = m.group(1).strip().rstrip("\n,.")
            candidates.append((m.start(), label))
        if len(candidates) >= 2:
            # Deduplicate
            filtered = []; last_pos = -999
            for pos, label in sorted(candidates):
                if pos - last_pos > 150:
                    filtered.append((pos, label)); last_pos = pos
            if len(filtered) >= 2:
                segments = []
                for i,(pos,label) in enumerate(filtered):
                    end = filtered[i+1][0] if i+1<len(filtered) else len(t)
                    segments.append({"label": label, "text": t[pos:end]})
                return segments

    # Strategy 3: Count repeated "Flow" / "Capacity" lines — if 3+, likely multi-pump
    flow_lines = re.findall(
        r"(?:Flow|Capacity|Discharge).*?\d+\.?\d*\s*m3",
        t, re.IGNORECASE
    )
    if len(flow_lines) >= 3:
        # Try to split on blank lines between sections
        chunks = re.split(r"\n{2,}", t)
        pump_chunks = [c for c in chunks if re.search(r"\d+\.?\d*\s*m3", c, re.IGNORECASE)]
        if len(pump_chunks) >= 2:
            segments = []
            for i, chunk in enumerate(pump_chunks):
                label = f"Pump {i+1}"
                # Try to extract a name from first line
                first_line = chunk.strip().split("\n")[0][:60]
                if len(first_line) > 5:
                    label = first_line.strip()
                segments.append({"label": label, "text": chunk})
            return segments

    return []


def parse_specs(text, learned_patterns=None):
    specs = {}
    t = text.replace("\n"," ").replace("  "," ")
    tl = t.lower()
    for field, patterns in _PATTERNS.items():
        for p in patterns:
            try:
                m = re.search(p, t)
                if not m: continue
                val = float(m.group(1))
                # Unit conversions
                if field == "flow_m3h":
                    if any(x in p for x in ["LPM", "L/[Mm]in"]):
                        val = round(val / 1000 * 60, 2)   # LPM → m³/h
                    elif any(x in p for x in ["[Ll][Pp][Ss]", "[Ll]/[Ss]", "l/s"]):
                        val = round(val * 3.6, 2)           # LPS → m³/h
                if field == "motor_kw" and "[Hh][Pp]" in p:
                    val = round(val * 0.7457, 2)            # HP → kW
                if field == "density_kgm3" and val < 5:
                    val = round(val * 1000, 1)              # SG → kg/m³
                lo, hi = _SANITY.get(field,(None,None))
                if lo is not None and not (lo < val < hi): continue
                specs[field] = val; break
            except: continue
    if learned_patterns:
        for lp in learned_patterns:
            fld=lp.get("field"); snip=lp.get("snippet","")
            if fld and snip and fld not in specs:
                try:
                    m = re.search(re.escape(" ".join(snip.split()[:3]))+r".*?(\d+\.?\d*)",
                                  t, re.IGNORECASE)
                    if m: specs[fld] = float(m.group(1))
                except: pass
    for kw, name in _FLUID_MAP:
        if kw in tl: specs["fluid"] = name; break
    if "fluid" not in specs: specs["fluid"] = "Clear Water"
    if "density_kgm3" not in specs:
        specs["density_kgm3"] = _DENS_DEF.get(specs["fluid"], 1000)
    for p in [
        r"[Pp]ump\s+[Mm]odel\s*[:\-=]?\s*([A-Za-z0-9][A-Za-z0-9\-\/\. ]{3,25})",
        r"[Mm]odel\s*[:\-\/=]?\s*([A-Z][A-Z0-9\-\/]{3,20})",
    ]:
        m = re.search(p, t)
        if m:
            val = m.group(1).strip().rstrip(".,;")
            if len(val)>=4: specs["model"]=val; break
    for mfr in ["Flowserve","KSB","Metso","Sulzer","Wilo","Jyoti",
                "Kirloskar","Grundfos","Ebara","Xylem","Andritz","Weir"]:
        if mfr.lower() in tl: specs["manufacturer"]=mfr; break
    return specs


# ═══════════════════════════════════════════════════════════════════
# SECTION 5 — TIER 1  (database match)
# ═══════════════════════════════════════════════════════════════════

def tier1_match(specs, db):
    pumps = db["pumps"].copy().dropna(subset=["Flow_m3h","Head_m"])
    Q     = specs.get("flow_m3h")
    H     = specs.get("head_m")
    model = (specs.get("model") or "").upper().strip()
    mfr   = (specs.get("manufacturer") or "").lower().strip()
    if not Q and not H and not model:
        return None, 0, "no_specs"
    best_row, best_score, best_type = None, 0, "none"
    for _, row in pumps.iterrows():
        score = 0
        db_model = str(row["Model"]).upper()
        db_mfr   = str(row["Manufacturer"]).lower()
        if model:
            m1 = re.sub(r"[\s\-\/]","",model)
            m2 = re.sub(r"[\s\-\/]","",db_model)
            if   m1==m2:              score+=55
            elif m1 in m2:            score+=50
            elif m2 in m1:            score+=45
            elif model[:6] in db_model: score+=25
        if mfr and db_mfr and mfr[:5] in db_mfr: score+=15
        if Q and pd.notna(row["Flow_m3h"]):
            pct = abs(Q-row["Flow_m3h"])/max(row["Flow_m3h"],1)
            score += 30 if pct<0.05 else 20 if pct<0.15 else 10 if pct<0.25 else 0
        if H and pd.notna(row["Head_m"]):
            pct = abs(H-row["Head_m"])/max(row["Head_m"],1)
            score += 25 if pct<0.05 else 15 if pct<0.15 else 8 if pct<0.20 else 0
        if score > best_score:
            best_score=score; best_row=row
            best_type = "exact" if score>=65 else "close" if score>=35 else "weak"
    return best_row, best_score, best_type

def get_bom_from_match(pump_row, db):
    model  = str(pump_row["Model"])
    pump_id= str(pump_row["Pump_ID"])
    matched = db["comps"][db["comps"]["Pump_Model_Compatibility"].str.contains(
        re.escape(model), case=False, na=False, regex=True
    )].copy()

    # Fill component weights from REAL_WEIGHTS where NaN
    if pump_id in REAL_WEIGHTS:
        wts = REAL_WEIGHTS[pump_id]
        # Map category → weight key
        cat_wt_map = {
            "Pump":      ["Pump (bare)", "Pump assembly", "Pump (VTP complete)",
                          "Pump (empty)", "Pump barrel + column"],
            "Motor":     ["Motor (CGL 550kW)", "Motor (Innomotics 500kW)",
                          "Motor (ABB 15kW)", "Motor (ABB 30kW)", "Motor (KECL 200kW)",
                          "Motor (190kW)", "Motor (315kW)", "Motor (160kW)",
                          "Motor (5.5kW)", "Motor (VTA)", "Motor (CGL VPC710 260kW)"],
            "Baseplate": ["Baseplate (IS2062)", "Bed Frame with Motor Support",
                          "Bed Frame with Motor", "Base frame + motor stool"],
            "Enclosure": ["Acoustic Enclosure (ARK)"],
            "Coupling":  ["Coupling (Rathi)", "V-Belt Guard"],
            "Rotor":     ["Rotor Assembly"],
            "Casing":    ["Top Casing Half"],
        }
        # Fill by category
        for idx2, row in matched.iterrows():
            if pd.notna(matched.at[idx2, "Weight_kg"]):
                continue
            cat = str(row.get("Category",""))
            keys = cat_wt_map.get(cat, [])
            for key in keys:
                if key in wts and isinstance(wts[key], (int,float)):
                    matched.at[idx2, "Weight_kg"] = wts[key]
                    break

    return matched


# ═══════════════════════════════════════════════════════════════════
# SECTION 6 — TIER 2  (physics)
# ═══════════════════════════════════════════════════════════════════

_IEC = [0.18,0.25,0.37,0.55,0.75,1.1,1.5,2.2,3,4,5.5,7.5,11,15,18.5,22,30,
        37,45,55,75,90,110,132,160,200,250,315,400,500,630,800,1000]

def round_iec(kw):
    for s in _IEC:
        if s>=kw: return s
    return round(kw,1)

def calc_specific_speed(Q, H, n=1450):
    if not Q or not H or Q<=0 or H<=0: return None
    return n * math.sqrt(Q*4.403) / ((H*3.281)**0.75)

def classify_pump_type(Ns, Q, H, fluid, stages=1, learned=None):
    fl = (fluid or "").lower()
    if any(k in fl for k in ["slurry","abrasive"]):
        return "TPL-HSS-01","Horizontal Slurry Pump"
    if any(k in fl for k in ["sulphuric","sulfuric","hydrochloric","acid"]):
        return "TPL-VSP-01","Vertical Sump Pump"
    if "live steam condensate" in fl:
        return "TPL-VTP-02","Vertical Turbine Pump VS6 (Condensate)"
    if "condensate" in fl:
        return "TPL-VTP-02","Vertical Turbine Pump VS6 (Condensate)"
    if "boiler feed" in fl:
        return "TPL-MSC-01","Multistage Centrifugal (BFW)"
    if learned: return "TPL-LEARNED", learned
    if Ns is None: return "TPL-HSC-01","Horizontal Split Casing (default)"
    if stages and stages>1 and H and H>150:
        return "TPL-MSC-01","Multistage Centrifugal"
    if Q and Q<20 and H and H>50:
        return "TPL-VRT-01","Vertical Submersible"
    if Ns < 1500:
        return ("TPL-HSC-02","Horizontal Split Casing — High Head") \
               if (H and H>150) else ("TPL-HSC-01","Horizontal Split Casing")
    elif Ns < 4000: return "TPL-VTP-01","Vertical Turbine Pump"
    else:           return "TPL-VTP-01","Vertical Turbine Pump (Axial)"

def calc_motor_kw(Q, H, rho=1000, eta=0.78, eta_m=0.93, sf=1.10):
    if not Q or not H or Q<=0 or H<=0: return None
    return round_iec((Q*H*rho*9.81)/(eta*3600*1000)/eta_m*sf)

def select_material(fluid, temp_c, pressure_kpa, db):
    mc = db["mat_compat"].copy()
    mc = mc.dropna(subset=["Rule_ID"])
    mc = mc[mc["Rule_ID"].astype(str).str.startswith("MAT-", na=False)]
    fl   = str(fluid or "").lower()
    temp = float(temp_c or 30)
    pres = float(pressure_kpa or 500)
    mc["Temp_Min_C"]       = pd.to_numeric(mc["Temp_Min_C"],       errors="coerce").fillna(0)
    mc["Temp_Max_C"]       = pd.to_numeric(mc["Temp_Max_C"],       errors="coerce").fillna(500)
    mc["Pressure_Max_kPa"] = pd.to_numeric(mc["Pressure_Max_kPa"],errors="coerce").fillna(99999)
    defaults = mc[mc["Rule_ID"].str.startswith("MAT-DEFAULT",na=False)]
    rules    = mc[~mc["Rule_ID"].str.startswith("MAT-DEFAULT",na=False)]
    valid = rules[
        (rules["Temp_Min_C"]<=temp) & (rules["Temp_Max_C"]>=temp) &
        (rules["Pressure_Max_kPa"]>=pres)
    ]
    exact = valid[valid["Fluid_Type"].str.lower().str.contains(fl[:12],na=False,case=False)]
    if not exact.empty:
        row = exact.iloc[0]
    else:
        cat_map = {"water":"Clean Water","caustic":"Alkali/Caustic",
                   "acid":"Dilute Acid","slurry":"Abrasive Slurry",
                   "condensate":"Condensate","oil":"Hydrocarbon",
                   "seawater":"Seawater","boiler":"High Temp"}
        cat = "Clean Water"
        for k,v in cat_map.items():
            if k in fl: cat=v; break
        cm = valid[valid["Fluid_Category"]==cat]
        row = cm.iloc[0] if not cm.empty \
              else (defaults.iloc[0] if not defaults.empty else rules.iloc[0])
    cols = ["Casing_MOC","Impeller_MOC","Shaft_MOC","Shaft_Sleeve_MOC",
            "Wear_Ring_MOC","Seal_Type","Seal_Plan","Fastener_MOC"]
    out = {c:(str(row[c]) if c in row.index and pd.notna(row[c]) else "VTA") for c in cols}
    out["Rule_ID"]     = str(row.get("Rule_ID",""))
    out["Fluid_Match"] = str(row.get("Fluid_Type",""))
    return out

def estimate_weight(pump_type_str, motor_kw, store=None):
    P  = float(motor_kw or 30)
    pt = (pump_type_str or "").lower()
    pc = mc = 1.0
    if store:
        for key, cal in store.get("weight_calibs",{}).items():
            if key in pt:
                pc = cal.get("pump_coeff",1.0)
                mc = cal.get("motor_coeff",1.0)
                break
    w = {}
    if "slurry" in pt:
        w["Pump (bare)"]    = round(_base_pump_wt(pt,P)*pc)
        w["Motor"]          = round(_base_motor_wt(P)*mc)
        w["Baseplate"]      = round(0.30*w["Pump (bare)"])
        w["Guard"]          = round(0.10*w["Pump (bare)"])
    elif "sump" in pt or "acid" in pt:
        w["Pump (bare)"]    = round(_base_pump_wt(pt,P)*pc)
        w["Motor"]          = round(_base_motor_wt(P)*mc)
        w["Baseplate"]      = 20
    elif "turbine" in pt or "vs6" in pt:
        w["Pump (bare)"]    = round(_base_pump_wt(pt,P)*pc)
        w["Motor"]          = round(_base_motor_wt(P)*mc)
        w["Baseplate"]      = 60
    else:
        w["Pump (bare)"]    = round(_base_pump_wt(pt,P)*pc)
        w["Rotor Assembly"] = round(w["Pump (bare)"]*0.15)
        w["Motor"]          = round(_base_motor_wt(P)*mc)
        w["Baseplate"]      = round(max(80, 0.18*P**1.02))
        w["Coupling"]       = max(10, round(0.015*P))
    w["TOTAL PACKAGE"] = sum(w.values())
    return w

def get_bom_template(template_id, db):
    sec_b = pd.read_excel(DB_PATH, sheet_name="BOM_Templates", header=16)
    sec_b.columns = [str(c).strip() for c in sec_b.columns]
    if "Template_ID" in sec_b.columns:
        tpl = sec_b[sec_b["Template_ID"].astype(str).str.strip()==template_id.strip()]
        if not tpl.empty: return tpl
    return sec_b[sec_b["Template_ID"].astype(str).str.contains("HSC-01",na=False)]

def tier2_generate(specs, db, store=None):
    Q=specs.get("flow_m3h"); H=specs.get("head_m")
    n=specs.get("speed_rpm") or 1450
    fluid=specs.get("fluid") or "Clear Water"
    temp=specs.get("temp_c") or 30
    rho=specs.get("density_kgm3") or 1000
    stages=specs.get("stages") or 1
    motor_input=specs.get("motor_kw")
    Ns  = calc_specific_speed(Q,H,n)
    lrn = get_learned_correction(Ns,fluid) if Ns else None
    tpl_id, pump_type = classify_pump_type(Ns,Q,H,fluid,stages,lrn)
    eta = 0.75 if (Ns and Ns<1500) else 0.82
    mkw = motor_input or calc_motor_kw(Q,H,rho,eta)
    moc = select_material(fluid,temp,500,db)
    wts = estimate_weight(pump_type, mkw, store)
    tpl = get_bom_template(tpl_id, db)
    moc_map = {
        "Casing":"Casing_MOC","Impeller":"Impeller_MOC","Shaft":"Shaft_MOC",
        "Sleeve":"Shaft_Sleeve_MOC","Wear Ring":"Wear_Ring_MOC",
        "Seal":"Seal_Type","Fasteners":"Fastener_MOC",
    }
    wt_map = {
        "Pump":"Pump (bare)","Motor":"Motor",
        "Baseplate":"Baseplate","Coupling":"Coupling",
    }
    rows = []
    for i,(_, tc) in enumerate(tpl.iterrows(), 1):
        cat = str(tc.get("Component_Category",""))
        sub = str(tc.get("Component_Subcategory",""))
        req = str(tc.get("Req_Type","M"))
        qty = str(tc.get("Qty_Logic","1"))
        mat = moc.get(moc_map.get(cat,"Casing_MOC"),"Per Service")
        if mat in ["nan","VTA","None",""]: mat="Engineer to Specify"
        h_info = HIERARCHY.get(cat, ("Z. OTHER", 3, "Other"))
        rows.append({
            "No":           i,
            "Component_ID": f"CALC-{tpl_id}-{i:03d}",
            "Category":     cat,
            "Sub_Assembly": h_info[2],
            "Description":  sub if sub and sub!=cat else cat,
            "MOC":          mat,
            "MOC_Rule":     moc.get("Rule_ID",""),
            "Qty":          qty,
            "Req_Type":     req,
            "Weight_kg":    wts.get(wt_map.get(cat,""),""),
            "Source":       "Tier 2 — Physics",
            "Notes":        str(tc.get("Notes_for_BOM_Generator",""))[:80],
        })
    summary = {
        "specific_speed_Ns":  round(Ns,1) if Ns else "N/A",
        "pump_type":          pump_type,
        "template_used":      tpl_id,
        "motor_kw_calc":      mkw,
        "eta_pump_assumed":   round(eta*100,1),
        "material_rule":      moc.get("Rule_ID",""),
        "fluid_matched":      moc.get("Fluid_Match",""),
        "seal_plan":          moc.get("Seal_Plan",""),
        "weights":            wts,
        "moc":                moc,
        "learned_correction": lrn,
    }
    return pd.DataFrame(rows), summary


# ═══════════════════════════════════════════════════════════════════
# SECTION 7 — HIERARCHICAL GROUPING
# ═══════════════════════════════════════════════════════════════════

def group_bom(bom_df):
    """
    Returns list of (section_name, sub_assembly, rows_df) tuples
    ordered by SECTION_ORDER.
    Adds hierarchy metadata: Section, Sub_Assembly, Level
    """
    if bom_df is None or bom_df.empty: return []
    cat_col = "Category" if "Category" in bom_df.columns else bom_df.columns[2]
    df = bom_df.copy()
    def _hier(cat):
        return HIERARCHY.get(str(cat).strip(), ("Z. OTHER", 3, "Other"))
    df["_section"]      = df[cat_col].apply(lambda c: _hier(c)[0])
    df["_level"]        = df[cat_col].apply(lambda c: _hier(c)[1])
    df["_sub_assembly"] = df[cat_col].apply(lambda c: _hier(c)[2])

    # Order by section
    sec_order = {s:i for i,s in enumerate(SECTION_ORDER)}
    df["_sec_ord"] = df["_section"].apply(lambda s: sec_order.get(s, 99))
    df = df.sort_values("_sec_ord").reset_index(drop=True)

    # Re-number
    df["No"] = range(1, len(df)+1)

    # Group
    result = []
    for sec in SECTION_ORDER:
        sec_rows = df[df["_section"]==sec].copy()
        if sec_rows.empty: continue
        # Further group by sub_assembly within section
        for sub in sec_rows["_sub_assembly"].unique():
            sub_rows = sec_rows[sec_rows["_sub_assembly"]==sub].copy()
            sub_rows = sub_rows.drop(columns=["_section","_level","_sub_assembly","_sec_ord"])
            result.append((sec, sub, sub_rows))

    # Ungrouped
    other = df[df["_section"]=="Z. OTHER"].copy()
    if not other.empty:
        other = other.drop(columns=["_section","_level","_sub_assembly","_sec_ord"])
        result.append(("Z. OTHER", "Other", other))

    return result


# ═══════════════════════════════════════════════════════════════════
# SECTION 8 — WEIGHT SCHEDULE BUILDER
# ═══════════════════════════════════════════════════════════════════

def build_weight_schedule(pump_id, tier, calc_summary=None):
    """
    Returns (weight_dict, crane_cat, foundation_dict)
    Uses real data for Tier 1, calculated for Tier 2.
    """
    if tier == "tier1" and pump_id in REAL_WEIGHTS:
        wts = REAL_WEIGHTS[pump_id]
    elif calc_summary and calc_summary.get("weights"):
        wts = calc_summary["weights"]
    else:
        wts = {}

    total = wts.get("TOTAL PACKAGE") or wts.get("total_kg") or sum(
        v for k,v in wts.items()
        if k not in ("TOTAL PACKAGE","total_kg","Heaviest Single Lift","Heaviest Lift Item")
        and isinstance(v,(int,float))
    )

    crane = crane_category(total)
    found = foundation_loads(total)

    return wts, crane, found


# ═══════════════════════════════════════════════════════════════════
# SECTION 9 — ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════════

def _safe(v):
    if v is None: return None
    if isinstance(v,str): return v.strip() or None
    try:
        f=float(v); return f if f==f else None
    except: return v

def generate_bom(specs, db, store=None):
    specs = {k:_safe(v) for k,v in (specs or {}).items()}
    pump_row, score, mtype = tier1_match(specs, db)
    if pump_row is not None and score>=30:
        bom = get_bom_from_match(pump_row, db)
        if not bom.empty:
            out = bom[["Component_ID","Category","Subcategory",
                        "Component_Name","Material_Spec","Qty_Per_Unit",
                        "Unit","Weight_kg","Vendor_Name","Notes"]].copy()
            # Add hierarchy columns
            def _hier(cat):
                return HIERARCHY.get(str(cat).strip(),("Z. OTHER",3,"Other"))
            out["Sub_Assembly"] = out["Category"].apply(lambda c: _hier(c)[2])
            out["MOC_Rule"]     = ""   # real data — rule implicit in Material_Spec
            out.insert(0,"No", range(1,len(out)+1))
            out.insert(9,"Source","Tier 1 — Database")
            return out, "tier1", {
                "pump_id":    pump_row["Pump_ID"],
                "model":      pump_row["Model"],
                "score":      score,
                "match_type": mtype,
            }, None
    bom_df, summary = tier2_generate(specs, db, store)
    return bom_df, "tier2", None, summary


# ═══════════════════════════════════════════════════════════════════
# SECTION 10 — EXCEL EXPORT  (4-tab professional output)
# ═══════════════════════════════════════════════════════════════════

def export_bom_excel(bom_df, specs, tier, match_info, calc_summary,
                     pump_id=None):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin = Side(style="thin",  color="CCCCCC")
    med  = Side(style="medium", color="1F4E79")
    bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)

    def hfill(color): return PatternFill("solid",start_color=color)
    def font(bold=False,size=9,color="000000"):
        return Font(bold=bold,size=size,color=color)
    def align(h="left",wrap=False):
        return Alignment(horizontal=h,vertical="center",wrap_text=wrap)

    # ── TAB 1: COVER PAGE ────────────────────────────────────────
    ws0 = wb.active; ws0.title = "Cover"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 35
    ws0.column_dimensions["B"].width = 55

    # Title block
    ws0.merge_cells("A1:B1")
    c = ws0["A1"]; c.value = "BILL OF MATERIALS"
    c.font = Font(bold=True,size=18,color="FFFFFF")
    c.fill = hfill("1F4E79")
    c.alignment = align("center"); ws0.row_dimensions[1].height=36

    ws0.merge_cells("A2:B2")
    c = ws0["A2"]
    model_str = (specs or {}).get("model","") or \
                (calc_summary or {}).get("pump_type","Generated BOM")
    c.value = model_str
    c.font = Font(bold=True,size=13,color="1F4E79")
    c.alignment = align("center"); ws0.row_dimensions[2].height=28

    ws0.merge_cells("A3:B3"); ws0.row_dimensions[3].height=10

    # Project info
    info = [
        ("BOM Method",
         f"Tier 1 — Database Match ({(match_info or {}).get('model','')})"
         if tier=="tier1"
         else f"Tier 2 — Physics Calculated ({(calc_summary or {}).get('pump_type','')})"),
        ("Flow",        f"{(specs or {}).get('flow_m3h','—')} m³/h"),
        ("Head",        f"{(specs or {}).get('head_m','—')} m"),
        ("Fluid",       (specs or {}).get('fluid','—')),
        ("Temperature", f"{(specs or {}).get('temp_c','—')} °C"),
        ("Motor",       f"{(specs or {}).get('motor_kw') or (calc_summary or {}).get('motor_kw_calc','—')} kW"),
        ("Generated",   pd.Timestamp.now().strftime("%d-%b-%Y %H:%M")),
    ]
    if tier=="tier1" and match_info:
        info += [
            ("Match Score", f"{match_info.get('score',0)}/100 ({match_info.get('match_type','')})"),
            ("Pump ID",     match_info.get("pump_id","")),
        ]
    if tier=="tier2" and calc_summary:
        info += [
            ("Specific Speed (Ns)", calc_summary.get("specific_speed_Ns","—")),
            ("Seal Plan",           calc_summary.get("seal_plan","—")),
            ("Material Rule",       calc_summary.get("material_rule","—")),
        ]

    r = 4
    for lbl, val in info:
        c1 = ws0.cell(r,1,lbl);     c1.font=font(True); c1.fill=hfill("EEF2F7"); c1.border=bdr; c1.alignment=align()
        c2 = ws0.cell(r,2,str(val)); c2.font=font();    c2.fill=hfill("F7F9FC"); c2.border=bdr; c2.alignment=align()
        ws0.row_dimensions[r].height=18; r+=1

    # ── TAB 2: HIERARCHICAL BOM ──────────────────────────────────
    ws1 = wb.create_sheet("BOM — Grouped")
    ws1.sheet_view.showGridLines = False
    col_w = [5,22,18,30,38,28,6,6,10,22,14,40]
    for i,w in enumerate(col_w):
        ws1.column_dimensions[get_column_letter(i+1)].width=w

    # Header
    ws1.merge_cells(f"A1:{get_column_letter(len(col_w))}1")
    c=ws1["A1"]; c.value="BILL OF MATERIALS — SUB-ASSEMBLY GROUPED VIEW"
    c.font=Font(bold=True,size=12,color="FFFFFF"); c.fill=hfill("1F4E79")
    c.alignment=align("center"); ws1.row_dimensions[1].height=24

    HEADERS = ["No","Component ID","Sub-Assembly","Category","Description / Name",
               "Material (MOC)","MOC Rule","Qty","Unit","Weight (kg)","Vendor","Notes"]
    r=2
    for j,h in enumerate(HEADERS):
        c=ws1.cell(r,j+1,h); c.font=font(True,9,"FFFFFF"); c.fill=hfill("2E75B6")
        c.alignment=align("center",True); c.border=bdr
    ws1.row_dimensions[r].height=26; r+=1

    ws1.freeze_panes="A3"

    SECTION_COLORS = {
        "A. PUMP HYDRAULICS":        "1A3A5C",
        "B. ROTATING ASSEMBLY":      "1A3A5C",
        "C. BEARINGS & LUBRICATION": "2E5984",
        "D. SHAFT SEALING":          "2E5984",
        "E. DRIVE & COUPLING":       "366092",
        "F. MOTOR / DRIVER":         "17375E",
        "G. STRUCTURAL":             "4F6228",
        "H. PIPING & NOZZLES":       "4F6228",
        "I. FASTENERS & GASKETS":    "595959",
        "J. INSTRUMENTATION":        "595959",
        "K. ACOUSTIC & SAFETY":      "7F7F7F",
        "L. COMPLETE ASSEMBLY":      "1F4E79",
        "Z. OTHER":                  "444444",
    }

    groups = group_bom(bom_df)
    current_section = None
    alt1 = hfill("EEF4FB"); alt2 = hfill("FFFFFF")
    row_count = 0

    for sec, sub, gdf in groups:
        # Section header (new section only)
        if sec != current_section:
            current_section = sec
            ws1.merge_cells(f"A{r}:{get_column_letter(len(col_w))}{r}")
            sc = ws1.cell(r,1,f"   {sec}")
            sc.font=Font(bold=True,size=9,color="FFFFFF")
            sc.fill=hfill(SECTION_COLORS.get(sec,"444444"))
            sc.alignment=align()
            ws1.row_dimensions[r].height=16; r+=1

        # Sub-assembly header (if more than one in section)
        ws1.merge_cells(f"A{r}:{get_column_letter(len(col_w))}{r}")
        sc2 = ws1.cell(r,1,f"      ▶  {sub}")
        sc2.font=Font(bold=True,size=8,color="1F4E79")
        sc2.fill=hfill("DCE6F1"); sc2.alignment=align()
        ws1.row_dimensions[r].height=14; r+=1

        is_t1 = "Component_Name" in gdf.columns
        for i,(_, row_s) in enumerate(gdf.iterrows()):
            rf = alt1 if row_count%2==0 else alt2
            row_count+=1
            def v(cols):
                for c in (cols if isinstance(cols,list) else [cols]):
                    if c in row_s.index and pd.notna(row_s[c]) and str(row_s[c]).strip():
                        return row_s[c]
                return ""
            vals=[
                v("No"),
                v("Component_ID"),
                v("Sub_Assembly") if "Sub_Assembly" in row_s.index else sub,
                v("Category"),
                v(["Component_Name","Description"]),
                v(["Material_Spec","MOC"]),
                v("MOC_Rule"),
                v(["Qty_Per_Unit","Qty"]),
                v("Unit"),
                v("Weight_kg"),
                v("Vendor_Name"),
                v("Notes"),
            ]
            for j,val in enumerate(vals):
                c=ws1.cell(r,j+1,val if pd.notna(val) else "")
                c.font=font(size=8); c.fill=rf; c.border=bdr
                c.alignment=align(wrap=True)
            ws1.row_dimensions[r].height=16; r+=1

    # ── TAB 3: WEIGHT SCHEDULE ───────────────────────────────────
    ws2 = wb.create_sheet("Weight Schedule")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width=40
    ws2.column_dimensions["B"].width=18
    ws2.column_dimensions["C"].width=35

    ws2.merge_cells("A1:C1")
    c=ws2["A1"]; c.value="WEIGHT SCHEDULE & LIFTING REQUIREMENTS"
    c.font=Font(bold=True,size=12,color="FFFFFF"); c.fill=hfill("1F4E79")
    c.alignment=align("center"); ws2.row_dimensions[1].height=24

    pid = pump_id or (match_info or {}).get("pump_id","")
    wts, crane, found = build_weight_schedule(pid, tier, calc_summary)

    r=2
    def ws_hdr(text, color="2E75B6"):
        ws2.merge_cells(f"A{r}:C{r}")
        c=ws2.cell(r,1,text); c.font=font(True,9,"FFFFFF")
        c.fill=hfill(color); c.alignment=align()
        ws2.row_dimensions[r].height=16

    def ws_row(label, val, note="", shade=False):
        bg = "F2F7FC" if shade else "FFFFFF"
        c1=ws2.cell(r,1,label); c1.font=font(True,9); c1.fill=hfill(bg); c1.border=bdr; c1.alignment=align()
        c2=ws2.cell(r,2,str(val)+" kg" if isinstance(val,(int,float)) else str(val))
        c2.font=font(size=9); c2.fill=hfill(bg); c2.border=bdr; c2.alignment=align("center")
        c3=ws2.cell(r,3,note); c3.font=font(size=8,color="595959"); c3.fill=hfill(bg); c3.border=bdr; c3.alignment=align(wrap=True)
        ws2.row_dimensions[r].height=16

    # Component weights
    ws_hdr("A. COMPONENT WEIGHT BREAKDOWN"); r+=1
    shade=False
    for k,v in wts.items():
        if k in ("Heaviest Single Lift","Heaviest Lift Item"): continue
        is_total = "TOTAL" in str(k).upper()
        note = "From GA drawing (actual)" if tier=="tier1" else "Empirical estimate ±20%"
        if is_total: note = "Sum of above"
        ws_row(k, v, note, shade)
        if is_total:
            ws2.cell(r-1,1).font=Font(bold=True,size=9)
            ws2.cell(r-1,2).font=Font(bold=True,size=9)
            ws2.cell(r-1,1).fill=hfill("D6E4F0")
            ws2.cell(r-1,2).fill=hfill("D6E4F0")
            ws2.cell(r-1,3).fill=hfill("D6E4F0")
        shade=not shade; r+=1
    r+=1

    # Heaviest single lift
    ws_hdr("B. HEAVIEST SINGLE LIFT", "4F6228"); r+=1
    hlift = wts.get("Heaviest Single Lift","—")
    hitem = wts.get("Heaviest Lift Item","See component breakdown")
    ws_row("Heaviest Item", hitem, ""); r+=1
    ws_row("Heaviest Single Lift Weight", hlift,
           "Crane must be rated for this + rigging factor 1.25"); r+=2

    # Crane requirement
    ws_hdr("C. CRANE SELECTION", "17375E"); r+=1
    total_kg = wts.get("TOTAL PACKAGE") or wts.get("total_kg",0)
    ws_row("Total Package Weight", total_kg, "Dry weight"); r+=1
    ws_row("Crane Requirement", crane, "Based on total package weight"); r+=1
    ws_row("Rigging Safety Factor", "1.25×",
           "Per IS 3938 / good engineering practice"); r+=2

    # Foundation loads
    ws_hdr("D. FOUNDATION LOADS", "366092"); r+=1
    ws_row("Static Load (total package)",   found["static_kg"],  "Dead weight on foundation"); r+=1
    ws_row("Dynamic Load (estimated)",      found["dynamic_kg"], found["note"]); r+=1
    ws_row("Recommended Foundation Type",
           "RCC block with non-shrink grout",
           "Fosroc Conbextra GP2 or equivalent"); r+=2

    # Notes
    ws_hdr("E. NOTES", "595959"); r+=1
    notes = [
        "All weights are dry weights unless stated.",
        "Operating weight = Dry weight + fluid fill weight.",
        f"Basis: {'Actual GA drawing data' if tier=='tier1' else 'Empirical correlations ±20%'}",
        "Civil structural engineer to design foundation for static + dynamic loads.",
        "Transport route survey required for packages > 10,000 kg.",
    ]
    for n in notes:
        ws2.merge_cells(f"A{r}:C{r}")
        c=ws2.cell(r,1,f"• {n}"); c.font=font(size=8,color="444444")
        ws2.row_dimensions[r].height=14; r+=1

    # ── TAB 4: MATERIAL TRACEABILITY ─────────────────────────────
    ws3 = wb.create_sheet("Material Traceability")
    ws3.sheet_view.showGridLines = False
    for i,w in enumerate([22,20,25,18,18,18,30]):
        ws3.column_dimensions[get_column_letter(i+1)].width=w

    ws3.merge_cells("A1:G1")
    c=ws3["A1"]; c.value="MATERIAL TRACEABILITY — MOC LINKED TO FLUID-TEMPERATURE RULES"
    c.font=Font(bold=True,size=12,color="FFFFFF"); c.fill=hfill("1F4E79")
    c.alignment=align("center"); ws3.row_dimensions[1].height=24

    hdrs=["Component","Category","Material (MOC)","Fluid Service",
          "Temp Range","Seal Plan","Traceability Rule"]
    r=2
    for j,h in enumerate(hdrs):
        c=ws3.cell(r,j+1,h); c.font=font(True,9,"FFFFFF")
        c.fill=hfill("2E75B6"); c.alignment=align("center"); c.border=bdr
    ws3.row_dimensions[r].height=22; r+=1

    # Build traceability rows
    moc_info = (calc_summary or {}).get("moc",{})
    fluid_match = (calc_summary or {}).get("fluid_matched","—")
    mat_rule    = (calc_summary or {}).get("material_rule","—")
    seal_plan   = (calc_summary or {}).get("seal_plan","—")
    temp_c      = (specs or {}).get("temp_c","—")

    shade=False
    alt1=hfill("EEF4FB"); alt2=hfill("FFFFFF")
    for _,row_s in bom_df.iterrows():
        cat = str(row_s.get("Category",""))
        is_t1 = "Component_Name" in row_s.index
        name  = str(row_s.get("Component_Name","") or row_s.get("Description",""))
        mat   = str(row_s.get("Material_Spec","") or row_s.get("MOC",""))
        rule  = str(row_s.get("MOC_Rule","") or mat_rule)
        rf    = alt1 if shade else alt2; shade=not shade

        # Determine seal plan per component
        sp = seal_plan if cat in ("Seal","Mechanical Seal","Gland") else "—"

        row_vals=[name[:45], cat, mat[:35], fluid_match[:25],
                  f"{temp_c}°C" if temp_c!="—" else "—",
                  sp, rule]
        for j,val in enumerate(row_vals):
            c=ws3.cell(r,j+1,str(val) if val else ""); c.font=font(size=8)
            c.fill=rf; c.border=bdr; c.alignment=align(wrap=True)
        ws3.row_dimensions[r].height=14; r+=1

    # MOC summary box
    if moc_info:
        r+=1
        ws3.merge_cells(f"A{r}:G{r}")
        c=ws3.cell(r,1,"MATERIAL SELECTION SUMMARY — FROM COMPATIBILITY MATRIX")
        c.font=font(True,9,"FFFFFF"); c.fill=hfill("2E5984")
        ws3.row_dimensions[r].height=16; r+=1
        moc_rows=[
            ("Fluid Service",    fluid_match),
            ("Casing MOC",       moc_info.get("Casing_MOC","—")),
            ("Impeller MOC",     moc_info.get("Impeller_MOC","—")),
            ("Shaft MOC",        moc_info.get("Shaft_MOC","—")),
            ("Shaft Sleeve MOC", moc_info.get("Shaft_Sleeve_MOC","—")),
            ("Wear Ring MOC",    moc_info.get("Wear_Ring_MOC","—")),
            ("Seal Plan",        moc_info.get("Seal_Plan","—")),
            ("Fastener MOC",     moc_info.get("Fastener_MOC","—")),
            ("Compatibility Rule", mat_rule),
        ]
        shade2=False
        for lbl,val in moc_rows:
            rf2=hfill("EEF4FB") if shade2 else hfill("FFFFFF"); shade2=not shade2
            c1=ws3.cell(r,1,lbl); c1.font=font(True,8); c1.fill=rf2; c1.border=bdr
            ws3.merge_cells(f"B{r}:G{r}")
            c2=ws3.cell(r,2,str(val)); c2.font=font(size=8); c2.fill=rf2; c2.border=bdr
            ws3.row_dimensions[r].height=14; r+=1

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf
