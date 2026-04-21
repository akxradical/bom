"""
Claude-Powered BOM Engine
═════════════════════════
Claude reads the datasheet. Claude generates the BOM.
Claude prices each component with live web search.

The old rule-based engine stays for:
  - Database matching (Tier 1 — 12 known pumps)
  - Weight schedule (REAL_WEIGHTS from dissection sheets)
  - Hierarchy grouping (SECTION_ORDER)

Everything else: Claude handles it.
Author: Ayush Kamle
"""

import json, re, time, os
import pandas as pd
from io import BytesIO

# ─────────────────────────────────────────────────────────────────
# ANTHROPIC CLIENT
# ─────────────────────────────────────────────────────────────────

_client = None

def _get_client():
    global _client
    if _client is None:
        import anthropic
        import streamlit as st
        key = st.secrets.get("ANTHROPIC_API_KEY", "")
        if not key:
            raise ValueError("ANTHROPIC_API_KEY not set in Streamlit secrets")
        _client = anthropic.Anthropic(api_key=key)
    return _client


def _call_claude(prompt, system="", use_search=False, max_tokens=4000):
    """Single Claude API call. Returns text response."""
    client = _get_client()
    kwargs = {
        "model":      "claude-sonnet-4-5",
        "max_tokens":  max_tokens,
        "messages":   [{"role": "user", "content": prompt}],
    }
    if system:
        kwargs["system"] = system
    if use_search:
        kwargs["tools"] = [{"type": "web_search_20250305", "name": "web_search"}]

    resp = client.messages.create(**kwargs)

    parts = []
    for block in resp.content:
        if hasattr(block, "text"):
            parts.append(block.text)
    return "\n".join(parts).strip()


def _parse_json(text):
    """Extract JSON from Claude response. Handles extra text, fences, trailing commas."""
    if not text:
        return None

    # Strip markdown fences
    clean = text.strip()
    clean = clean.replace("```json", "").replace("```", "").strip()

    # Strategy 1: direct parse
    try:
        return json.loads(clean)
    except Exception:
        pass

    # Strategy 2: find the outermost [ ] or { } using bracket counting
    result = _bracket_extract(clean, '[', ']')
    if result is not None:
        return result

    result = _bracket_extract(clean, '{', '}')
    if result is not None:
        return result

    return None


def _bracket_extract(text, opener, closer):
    """Find and parse the first complete JSON structure bounded by opener/closer."""
    start = text.find(opener)
    if start == -1:
        return None

    depth = 0
    in_string = False
    i = start
    while i < len(text):
        ch = text[i]

        # Handle string literals — skip everything inside quotes
        if ch == '"' and (i == 0 or text[i-1] != '\\'):
            in_string = not in_string
            i += 1
            continue

        if in_string:
            i += 1
            continue

        if ch == opener:
            depth += 1
        elif ch == closer:
            depth -= 1
            if depth == 0:
                candidate = text[start:i+1]
                # Try parsing
                try:
                    return json.loads(candidate)
                except json.JSONDecodeError:
                    # Fix trailing commas: ,} or ,]
                    import re as _re
                    fixed = _re.sub(r',\s*([}\]])', r'\1', candidate)
                    try:
                        return json.loads(fixed)
                    except Exception:
                        return None
        i += 1

    return None


# ═══════════════════════════════════════════════════════════════════
# STEP 1 — CLAUDE READS THE PDF
# ═══════════════════════════════════════════════════════════════════

SPEC_SYSTEM = """You are a senior mechanical/procurement engineer specialising in 
rotating equipment (pumps, compressors, turbines) for EPC projects in India.

Your job: Read pump datasheets, GA drawings, vendor technical documents, and 
procurement specifications. Extract every technical parameter accurately.

Critical rules:
- If the document contains multiple pump types (e.g. Hydrant, Spray, Jockey), 
  identify ALL of them and list each separately.
- Convert all units to SI: flow in m³/h, head in metres, power in kW, temp in °C.
- If head is in mWC, mlc, or kPa — convert to metres.
- If flow is in LPM, LPS, USGPM, l/s — convert to m³/h.
- If power is in HP/BHP — convert to kW (×0.7457).
- Extract Material of Construction (MOC) for every component mentioned.
- Identify the pump type: HSC, VTP, Slurry, Sump, Multistage, etc.
- Note manufacturer, model, tag numbers, project details.
- If a field says "VTA" or "Bidder to furnish" → mark as null, don't guess."""


def claude_extract_specs(pdf_text):
    """
    Send PDF text to Claude. Returns structured specs dict.
    Handles multi-pump documents automatically.
    """
    prompt = f"""Read this pump technical document and extract ALL pump specifications.

DOCUMENT TEXT:
{pdf_text[:12000]}

Respond with ONLY a JSON object (no other text):
{{
  "document_type": "vendor_datasheet | procurement_spec | ga_drawing | manual",
  "project": "project name if mentioned",
  "manufacturer": "manufacturer name or null",
  "multi_pump": true/false,
  
  "pumps": [
    {{
      "pump_label": "descriptive name e.g. Hydrant Water Pump",
      "model": "pump model or null",
      "manufacturer": "name or null",
      "type": "Horizontal Split Casing | Vertical Turbine | Slurry | Sump | Multistage | Other",
      "tag_numbers": "tag nos or null",
      "standard": "API 610 | IS 5120 | etc or null",
      
      "flow_m3h": number or null,
      "head_m": number or null,
      "speed_rpm": number or null,
      "motor_kw": number or null,
      "stages": number or null,
      "temp_c": number or null,
      "density_kgm3": number or null,
      "fluid": "fluid name",
      "viscosity": "value or null",
      "npsha_m": number or null,
      
      "moc": {{
        "casing": "material spec or null",
        "impeller": "material spec or null",
        "shaft": "material spec or null",
        "shaft_sleeve": "material spec or null",
        "wear_ring": "material spec or null",
        "bearing": "material spec or null",
        "seal_type": "mechanical seal / gland packing / null",
        "seal_plan": "Plan 11, Plan 53B, etc or null",
        "baseplate": "material spec or null",
        "fasteners": "material spec or null"
      }},
      
      "nozzles": {{
        "suction_size": "size or null",
        "discharge_size": "size or null",
        "rating": "PN16, 150#, etc or null"
      }},

      "weights": {{
        "pump_bare_kg": number or null,
        "motor_kg": number or null,
        "baseplate_kg": number or null,
        "total_package_kg": number or null
      }},
      
      "coupling": "type or null",
      "drive": "motor | diesel engine | both",
      "notes": "any critical notes"
    }}
  ]
}}

Be accurate. If data is missing, use null — never guess."""

    raw = _call_claude(prompt, system=SPEC_SYSTEM)
    data = _parse_json(raw)
    return data


# ═══════════════════════════════════════════════════════════════════
# STEP 2 — CLAUDE GENERATES THE BOM
# ═══════════════════════════════════════════════════════════════════

BOM_SYSTEM = """You are an expert pump procurement engineer in India. 
You generate Bills of Materials for centrifugal pump packages.

Every BOM you generate must:
1. Cover ALL sub-assemblies: pump hydraulics, rotating assembly, bearings, 
   sealing, drive/coupling, motor, structural, piping/nozzles, fasteners, 
   instrumentation, acoustic (if needed), complete assembly.
2. Specify MOC (Material of Construction) for EVERY component.
3. Include realistic weights where known.
4. Specify quantity per pump unit.
5. Mark each component as M (Mandatory), C (Conditional), or O (Optional).
6. Be specific — not generic. E.g. "SKF 6217 C3" not just "Bearing".
7. Include seal plan piping, coupling guard, foundation bolts, counter flanges.
8. For Indian EPC: include RTDs for pump bearings, dial thermometers.
"""


def claude_generate_bom(pump_specs):
    """
    Given extracted pump specs, generate a complete BOM.
    Returns a list of component dicts.
    """
    specs_str = json.dumps(pump_specs, indent=2, default=str)

    prompt = f"""Generate a COMPLETE Bill of Materials for this pump:

PUMP SPECIFICATIONS:
{specs_str}

Generate the BOM as a JSON array. Each component:
{{
  "section": "A. PUMP HYDRAULICS | B. ROTATING ASSEMBLY | C. BEARINGS & LUBRICATION | D. SHAFT SEALING | E. DRIVE & COUPLING | F. MOTOR / DRIVER | G. STRUCTURAL | H. PIPING & NOZZLES | I. FASTENERS & GASKETS | J. INSTRUMENTATION | K. ACOUSTIC & SAFETY | L. COMPLETE ASSEMBLY",
  "sub_assembly": "Casing Assembly | Rotor Assembly | Bearing Assembly | Sealing Assembly | Drive Assembly | Motor | Structural | Piping Assembly | Fasteners | Instrumentation | Acoustic Enclosure | Complete Package",
  "component": "specific component name",
  "description": "detailed description with specs where applicable",
  "moc": "exact material specification — ASTM/IS/EN standard",
  "qty": "1 | 2 | 1 set | etc",
  "weight_kg": number or null,
  "req_type": "M | C | O",
  "notes": "any relevant notes"
}}

IMPORTANT:
- Include AT LEAST 25 components for a standard pump
- Include ALL wetted parts with correct MOC for the fluid service
- Motor: specify frame size, kW, voltage, poles
- Seal: specify type, plan, materials
- Coupling: specify type (disc/tyre/spacer), DBSE
- Baseplate: IS 2062 fabricated with drain pan
- Foundation bolts: specify quantity and size
- Counter flanges: both suction and discharge
- Gaskets: specify type and material
- Instrumentation: RTDs for pump bearings if applicable

Respond with ONLY the JSON array."""

    raw = _call_claude(prompt, system=BOM_SYSTEM, max_tokens=6000)
    data = _parse_json(raw)

    if isinstance(data, list):
        return data
    elif isinstance(data, dict) and "components" in data:
        return data["components"]
    elif isinstance(data, dict) and "bom" in data:
        return data["bom"]
    return data


def bom_to_dataframe(bom_list):
    """Convert Claude BOM output to a clean DataFrame."""
    if not bom_list:
        return pd.DataFrame()

    rows = []
    for i, comp in enumerate(bom_list, 1):
        rows.append({
            "No":           i,
            "Section":      str(comp.get("section", "")),
            "Sub_Assembly": str(comp.get("sub_assembly", "")),
            "Component":    str(comp.get("component", "")),
            "Description":  str(comp.get("description", "")),
            "MOC":          str(comp.get("moc", "")),
            "Qty":          str(comp.get("qty", "1")),
            "Weight_kg":    comp.get("weight_kg"),
            "Req_Type":     str(comp.get("req_type", "M")),
            "Notes":        str(comp.get("notes", "")),
        })
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════
# STEP 3 — CLAUDE PRICES EACH COMPONENT (with web search)
# ═══════════════════════════════════════════════════════════════════

PRICE_SYSTEM = """You are a procurement cost estimator for an EPC company in India.
You estimate current 2025-2026 market prices for pump components.

Rules:
- Prices in Indian Rupees (INR)
- Use current Indian market rates
- For castings: price depends on material and weight
- For motors: price depends on kW rating and voltage
- For seals: price depends on type, size, and plan
- Search for current vendor prices when possible
- Be realistic — not textbook prices, actual procurement prices
- Include machining, finishing, and quality testing costs
- For imported items, include customs + freight to India site"""


def claude_price_bom(bom_df, pump_specs, progress_callback=None):
    """
    Price every component in the BOM using Claude + web search.
    Groups similar items to reduce API calls.
    Returns enhanced DataFrame with price columns.
    """
    if bom_df is None or bom_df.empty:
        return bom_df

    specs_str = json.dumps(pump_specs, indent=2, default=str) if isinstance(pump_specs, dict) else str(pump_specs)

    # Build component list for batch pricing
    components = []
    for _, row in bom_df.iterrows():
        components.append({
            "no":          row.get("No", ""),
            "component":   str(row.get("Component", "")),
            "description": str(row.get("Description", "")),
            "moc":         str(row.get("MOC", "")),
            "qty":         str(row.get("Qty", "1")),
            "weight_kg":   row.get("Weight_kg"),
            "section":     str(row.get("Section", "")),
        })

    # Split into batches of ~10 for manageable API calls
    batch_size = 10
    batches    = [components[i:i+batch_size]
                  for i in range(0, len(components), batch_size)]

    all_prices = []
    progress_labels = [
        "Accumulating pump assembly market data...",
        "Compiling rotating equipment pricing indices...",
        "Gathering bearing and seal vendor rates...",
        "Analysing structural component costs...",
        "Finalising piping and accessories pricing...",
        "Cross-referencing vendor quotations...",
        "Validating price consistency...",
    ]

    for batch_idx, batch in enumerate(batches):
        if progress_callback:
            pct = int(10 + (batch_idx / max(len(batches), 1)) * 80)
            label = progress_labels[min(batch_idx, len(progress_labels)-1)]
            progress_callback(pct, label)

        batch_str = json.dumps(batch, indent=1, default=str)
        prompt = f"""Price these pump components at current Indian market rates (2025-2026).

PUMP CONTEXT:
{specs_str[:2000]}

COMPONENTS TO PRICE:
{batch_str}

For EACH component, respond with a JSON array:
[
  {{
    "no": <component number>,
    "unit_price_inr": <integer price in INR>,
    "total_price_inr": <unit_price × quantity>,
    "price_basis": "per unit | per kg | per set",
    "confidence": "high | medium | low",
    "source": "vendor/market source",
    "notes": "price basis, assumptions"
  }}
]

Pricing guidelines:
- Motor: ₹4,000-8,000 per kW for LT, ₹3,500-5,500 per kW for HT
- CS castings (A216 WCB): ₹200-280/kg finished
- SS316 castings: ₹700-900/kg finished
- High chrome (A532): ₹900-1200/kg finished
- Mechanical seal (cartridge): ₹40,000-4,00,000 depending on size/plan
- Bearings: ₹2,000-50,000 depending on type/size
- Coupling (disc/tyre): ₹15,000-80,000 depending on torque rating
- Baseplate IS2062: ₹100-140/kg fabricated
- Counter flanges: ₹1,500-8,000 depending on size/rating
- Foundation bolts: ₹150-400 per bolt depending on size
- Gaskets: ₹300-1,500 each depending on size/material

Use web search for motors, seals, and any component > ₹50,000.
Respond with ONLY the JSON array."""

        raw = _call_claude(prompt, system=PRICE_SYSTEM,
                          use_search=True, max_tokens=3000)
        data = _parse_json(raw)

        if isinstance(data, list):
            all_prices.extend(data)
        elif isinstance(data, dict):
            all_prices.append(data)

    # Merge prices into BOM
    if progress_callback:
        progress_callback(92, "Compiling final cost report...")

    price_map = {}
    for p in all_prices:
        if isinstance(p, dict) and "no" in p:
            price_map[p["no"]] = p

    result_rows = []
    for _, row in bom_df.iterrows():
        rd = row.to_dict()
        no = row.get("No", 0)
        pr = price_map.get(no, {})

        unit_p  = int(pr.get("unit_price_inr", 0))
        total_p = int(pr.get("total_price_inr", unit_p))
        gst     = int(total_p * 0.18)

        rd["Unit_Price_INR"]   = unit_p
        rd["Total_Price_INR"]  = total_p
        rd["GST_18pct"]        = gst
        rd["Price_With_GST"]   = total_p + gst
        rd["Price_Confidence"] = str(pr.get("confidence", "low"))
        rd["Price_Source"]     = str(pr.get("source", "estimate"))
        rd["Price_Notes"]      = str(pr.get("notes", ""))
        result_rows.append(rd)

    return pd.DataFrame(result_rows)


def build_cost_summary(priced_df):
    """Build a cost summary from priced BOM."""
    if priced_df is None or priced_df.empty:
        return {}

    total_ex  = int(priced_df["Total_Price_INR"].sum())
    total_gst = int(priced_df["GST_18pct"].sum())
    total_inc = int(priced_df["Price_With_GST"].sum())

    sub_col = "Sub_Assembly" if "Sub_Assembly" in priced_df.columns else "Section"
    sub_totals = (priced_df.groupby(sub_col)["Total_Price_INR"]
                  .sum().sort_values(ascending=False).to_dict())

    top5 = (priced_df.nlargest(5, "Total_Price_INR")
            [["Component","Description","Total_Price_INR","Price_Confidence"]]
            .to_dict("records"))

    conf = priced_df["Price_Confidence"].value_counts().to_dict()

    return {
        "total_ex_gst":   total_ex,
        "total_gst":      total_gst,
        "total_incl_gst": total_inc,
        "sub_totals":     {k: int(v) for k, v in sub_totals.items()},
        "top5_drivers":   top5,
        "confidence":     conf,
        "component_count":len(priced_df),
        "note": ("Indicative market estimate for budget planning. "
                 "Actual prices subject to vendor quotation."),
    }


# ═══════════════════════════════════════════════════════════════════
# GROUPING (for display — uses section from Claude output)
# ═══════════════════════════════════════════════════════════════════

SECTION_ORDER = [
    "A. PUMP HYDRAULICS",
    "B. ROTATING ASSEMBLY",
    "C. BEARINGS & LUBRICATION",
    "D. SHAFT SEALING",
    "E. DRIVE & COUPLING",
    "F. MOTOR / DRIVER",
    "G. STRUCTURAL",
    "H. PIPING & NOZZLES",
    "I. FASTENERS & GASKETS",
    "J. INSTRUMENTATION",
    "K. ACOUSTIC & SAFETY",
    "L. COMPLETE ASSEMBLY",
]

def group_bom(bom_df):
    """Group BOM by section and sub-assembly. Returns [(section, sub, df)]."""
    if bom_df is None or bom_df.empty:
        return []

    sec_col = "Section" if "Section" in bom_df.columns else None
    sub_col = "Sub_Assembly" if "Sub_Assembly" in bom_df.columns else None

    if not sec_col:
        return [("ALL", "All Components", bom_df)]

    sec_order = {s: i for i, s in enumerate(SECTION_ORDER)}
    df = bom_df.copy()
    df["_ord"] = df[sec_col].apply(lambda s: sec_order.get(str(s).strip(), 99))
    df = df.sort_values("_ord").reset_index(drop=True)
    df["No"] = range(1, len(df)+1)

    result = []
    for sec in SECTION_ORDER:
        sec_rows = df[df[sec_col].str.strip() == sec]
        if sec_rows.empty:
            continue
        if sub_col:
            for sub in sec_rows[sub_col].unique():
                sub_rows = sec_rows[sec_rows[sub_col] == sub].copy()
                sub_rows = sub_rows.drop(columns=["_ord"], errors="ignore")
                result.append((sec, str(sub), sub_rows))
        else:
            sec_rows2 = sec_rows.drop(columns=["_ord"], errors="ignore")
            result.append((sec, sec, sec_rows2))

    # Ungrouped
    other = df[~df[sec_col].str.strip().isin(SECTION_ORDER)]
    if not other.empty:
        other = other.drop(columns=["_ord"], errors="ignore")
        result.append(("Z. OTHER", "Other", other))

    return result


# ═══════════════════════════════════════════════════════════════════
# PDF EXTRACTION (still pdfplumber — free, local)
# ═══════════════════════════════════════════════════════════════════

def extract_pdf_text(file_bytes):
    """Extract text from PDF using pdfplumber."""
    try:
        import pdfplumber
        pages = []
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            for p in pdf.pages:
                t = p.extract_text()
                if t:
                    pages.append(t)
        return "\n".join(pages), None
    except Exception as e:
        return "", str(e)


# ═══════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════

def export_excel(bom_df, pump_specs, priced=False):
    """Export BOM to professional Excel with cover + grouped BOM + optional pricing."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb  = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfill(c): return PatternFill("solid", start_color=c)
    def hfont(b=False, s=9, c="000000"): return Font(bold=b, size=s, color=c)

    # ── Cover ─────────────────────────────────────────────────────
    ws0 = wb.active; ws0.title = "Cover"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 30; ws0.column_dimensions["B"].width = 55

    ws0.merge_cells("A1:B1")
    c = ws0["A1"]; c.value = "BILL OF MATERIALS"
    c.font = Font(bold=True, size=18, color="FFFFFF"); c.fill = hfill("1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws0.row_dimensions[1].height = 36

    specs = pump_specs if isinstance(pump_specs, dict) else {}
    info_items = []
    if isinstance(specs.get("pumps"), list) and specs["pumps"]:
        p = specs["pumps"][0]
        info_items = [
            ("Pump Model",       p.get("model", "—")),
            ("Manufacturer",     p.get("manufacturer", "—")),
            ("Type",             p.get("type", "—")),
            ("Flow (m³/h)",      p.get("flow_m3h", "—")),
            ("Head (m)",         p.get("head_m", "—")),
            ("Motor (kW)",       p.get("motor_kw", "—")),
            ("Fluid",            p.get("fluid", "—")),
            ("Temperature (°C)", p.get("temp_c", "—")),
            ("Standard",         p.get("standard", "—")),
            ("Project",          specs.get("project", "—")),
        ]
    else:
        info_items = [
            ("Pump",    specs.get("pump_label", "—")),
            ("Flow",    specs.get("flow_m3h", "—")),
            ("Head",    specs.get("head_m", "—")),
            ("Motor",   specs.get("motor_kw", "—")),
            ("Fluid",   specs.get("fluid", "—")),
        ]

    r = 3
    for lbl, val in info_items:
        ws0.cell(r, 1, lbl).font = hfont(True); ws0.cell(r, 1).fill = hfill("EEF2F7")
        ws0.cell(r, 1).border = bdr
        ws0.cell(r, 2, str(val) if val else "—").font = hfont()
        ws0.cell(r, 2).border = bdr
        r += 1

    ws0.cell(r+1, 1, "Generated").font = hfont(True)
    ws0.cell(r+1, 2, pd.Timestamp.now().strftime("%d-%b-%Y %H:%M")).font = hfont()

    # ── BOM Sheet ─────────────────────────────────────────────────
    ws1 = wb.create_sheet("BOM")
    ws1.sheet_view.showGridLines = False

    # Determine columns
    if priced and "Total_Price_INR" in bom_df.columns:
        cols = ["No","Section","Sub_Assembly","Component","Description","MOC",
                "Qty","Weight_kg","Req_Type","Unit_Price_INR","Total_Price_INR",
                "GST_18pct","Price_With_GST","Price_Confidence","Notes"]
    else:
        cols = ["No","Section","Sub_Assembly","Component","Description","MOC",
                "Qty","Weight_kg","Req_Type","Notes"]
    cols = [c for c in cols if c in bom_df.columns]

    widths = {"No":5,"Section":22,"Sub_Assembly":20,"Component":30,
              "Description":40,"MOC":25,"Qty":8,"Weight_kg":10,
              "Req_Type":6,"Unit_Price_INR":14,"Total_Price_INR":14,
              "GST_18pct":12,"Price_With_GST":14,"Price_Confidence":10,"Notes":35}

    ws1.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    c1 = ws1["A1"]; c1.value = "BILL OF MATERIALS"
    c1.font = Font(bold=True, size=12, color="FFFFFF"); c1.fill = hfill("1F4E79")
    c1.alignment = Alignment(horizontal="center"); ws1.row_dimensions[1].height = 24

    r = 2
    for j, col in enumerate(cols):
        c = ws1.cell(r, j+1, col.replace("_"," "))
        c.font = hfont(True, 9, "FFFFFF"); c.fill = hfill("2E75B6")
        c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = bdr
        ws1.column_dimensions[get_column_letter(j+1)].width = widths.get(col, 14)
    ws1.row_dimensions[r].height = 26; r += 1

    alt1 = hfill("EEF4FB"); alt2 = hfill("FFFFFF")
    for i, (_, row) in enumerate(bom_df.iterrows()):
        rf = alt1 if i % 2 == 0 else alt2
        for j, col in enumerate(cols):
            val = row.get(col, "")
            if pd.isna(val): val = ""
            if col in ("Unit_Price_INR","Total_Price_INR","GST_18pct","Price_With_GST"):
                if val and val != "": val = f"₹{int(val):,}"
            c = ws1.cell(r, j+1, val)
            c.font = hfont(size=8); c.fill = rf; c.border = bdr
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws1.row_dimensions[r].height = 16; r += 1

    ws1.freeze_panes = "A3"

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf
