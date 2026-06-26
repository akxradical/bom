"""
Agentic BOM + Should-Cost Engine
════════════════════════════════════════════════════════════════════
Universal engineered-product procurement engine.

Drop ANY engineered-product datasheet (pump, compressor, agitator, motor,
valve, fan/blower, heat exchanger, gearbox, conveyor, crane/MHE, HVAC
chiller, pressure vessel, turbine, ...). The agent:

  1. IDENTIFY   — figures out what the product is
  2. SCHEMA     — builds a product-specific sub-assembly schema from scratch
  3. BOM        — populates components for every sub-assembly
  4. VALIDATE   — checks completeness, loops to fill gaps (agentic)
  5. PRICE      — should-cost from live market rates (RM + machining, floor)
  6. CONFIDENCE — scores the run

No pump-specific hardcoding. No fixed A–L sections. The agent decides the
schema dynamically, every time, for every product.

Author: Ayush Kamle
"""

import json, re, time
import pandas as pd
from io import BytesIO

# ═══════════════════════════════════════════════════════════════════
# TOKEN USAGE TRACKING (live counter + ₹ cost)
# ═══════════════════════════════════════════════════════════════════

_USAGE = {"input": 0, "output": 0, "cached": 0, "calls": 0, "model": ""}

# ₹ per 1M tokens (input, output) — approx, ₹85/$. Picked by model family.
_RATES_INR = {
    "haiku": (85, 425), "sonnet": (255, 1275), "opus": (1275, 6375),
    "fable": (255, 1275), "free": (0, 0),
}

def _reset_usage():
    _USAGE.update(input=0, output=0, cached=0, calls=0, model="")

def _rate_key(model):
    m = (model or "").lower()
    for k in ("haiku", "sonnet", "opus", "fable"):
        if k in m:
            return k
    return "free"

def _track(inp, out, cached=0, model="free"):
    _USAGE["input"] += int(inp or 0)
    _USAGE["output"] += int(out or 0)
    _USAGE["cached"] += int(cached or 0)
    _USAGE["calls"] += 1
    if model and model != "free":
        _USAGE["model"] = model

def _est_tokens(*texts):
    return sum(len(str(t)) for t in texts) // 4   # ~4 chars/token

def usage_cost_inr():
    rk = _rate_key(_USAGE["model"])
    ri, ro = _RATES_INR.get(rk, (0, 0))
    return round(_USAGE["input"] / 1e6 * ri + _USAGE["output"] / 1e6 * ro, 2)

def usage_snapshot():
    return {**_USAGE, "total_tokens": _USAGE["input"] + _USAGE["output"],
            "est_cost_inr": usage_cost_inr()}


# ═══════════════════════════════════════════════════════════════════
# LLM PROVIDERS — Claude primary, free LLMs as fallback
# ═══════════════════════════════════════════════════════════════════

def _get_key(name):
    try:
        import streamlit as st
        v = st.secrets.get(name, "")
        if v: return v
    except Exception:
        pass
    import os
    return os.environ.get(name, "")


def _http_post(url, headers, body, timeout=90, retries=2):
    import urllib.request, urllib.error
    raw = json.dumps(body).encode()
    for i in range(retries):
        try:
            req = urllib.request.Request(url, data=raw, headers=headers)
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return json.loads(r.read())
        except urllib.error.HTTPError as e:
            if e.code in (429, 500, 502, 503, 504) and i < retries - 1:
                time.sleep(10 * (i + 1)); continue
            try: b = e.read().decode()[:300]
            except: b = str(e.reason)
            raise Exception(f"HTTP {e.code}: {b}")
        except Exception as e:
            if i < retries - 1: time.sleep(5); continue
            raise


def _oai(url, key, model, prompt, system="", mt=4000):
    msgs = []
    if system: msgs.append({"role": "system", "content": system})
    msgs.append({"role": "user", "content": prompt})
    d = _http_post(url, {"Content-Type": "application/json",
        "Authorization": f"Bearer {key}"},
        {"model": model, "messages": msgs, "max_tokens": mt, "temperature": 0.1})
    out = d["choices"][0]["message"]["content"].strip()
    u = d.get("usage") or {}
    _track(u.get("prompt_tokens") or _est_tokens(system, prompt),
           u.get("completion_tokens") or _est_tokens(out), 0, "free")
    return out


def _gemini_parse(d):
    """Defensively pull text from a Gemini response. Raises an informative
    error instead of KeyError when the model returns no text (e.g. the 2.5
    thinking model spends its whole budget on thoughts, or content is blocked)."""
    cands = d.get("candidates") or []
    if not cands:
        fb = d.get("promptFeedback", {})
        raise Exception(f"no candidates (blocked? {fb})")
    cand = cands[0]
    parts = (cand.get("content") or {}).get("parts") or []
    text = " ".join(p.get("text", "") for p in parts if "text" in p).strip()
    if not text:
        raise Exception(f"empty text (finishReason={cand.get('finishReason')})")
    um = d.get("usageMetadata") or {}
    _track(um.get("promptTokenCount") or _est_tokens(text) * 3,
           um.get("candidatesTokenCount") or _est_tokens(text), 0, "free")
    return text


# gemini-2.0-flash is GA, free-tier, fast, and (unlike 2.5-flash) does NOT burn
# the output budget on hidden 'thinking' tokens — so it reliably returns text.
_GEMINI_MODEL = "gemini-2.0-flash"


def _gemini(p, s="", mt=8000):
    k = _get_key("GEMINI_API_KEY")
    if not k: raise ValueError("no key")
    b = {"contents": [{"parts": [{"text": p}]}],
         "generationConfig": {"maxOutputTokens": mt, "temperature": 0.1}}
    if s: b["systemInstruction"] = {"parts": [{"text": s}]}
    d = _http_post(f"https://generativelanguage.googleapis.com/v1beta/models/{_GEMINI_MODEL}:generateContent?key={k}",
                   {"Content-Type": "application/json"}, b)
    return _gemini_parse(d)


def _gemini_grounded(p, s="", mt=4000):
    """Gemini with Google Search grounding — the free equivalent of Claude
    web_search. The model reads live Google results and answers from them,
    so prices come from real pages, not hallucination."""
    k = _get_key("GEMINI_API_KEY")
    if not k: raise ValueError("GEMINI_API_KEY not set")
    b = {"contents": [{"parts": [{"text": p}]}],
         "tools": [{"google_search": {}}],
         "generationConfig": {"maxOutputTokens": mt, "temperature": 0.1}}
    if s: b["systemInstruction"] = {"parts": [{"text": s}]}
    d = _http_post(f"https://generativelanguage.googleapis.com/v1beta/models/{_GEMINI_MODEL}:generateContent?key={k}",
                   {"Content-Type": "application/json"}, b, timeout=120)
    return _gemini_parse(d)


def _groq(p, s="", mt=4000):
    k = _get_key("GROQ_API_KEY")
    if not k: raise ValueError("no key")
    last = "unknown"
    for m in ["llama-3.3-70b-versatile", "llama-3.1-8b-instant"]:
        try: return _oai("https://api.groq.com/openai/v1/chat/completions", k, m, p, s, mt)
        except Exception as e: last = str(e)
    raise Exception(f"Groq failed: {last[:100]}")


def _mistral(p, s="", mt=4000):
    k = _get_key("MISTRAL_API_KEY")
    if not k: raise ValueError("no key")
    return _oai("https://api.mistral.ai/v1/chat/completions", k, "mistral-large-latest", p, s, mt)


def _openrouter(p, s="", mt=4000):
    k = _get_key("OPENROUTER_API_KEY")
    if not k: raise ValueError("no key")
    last = "unknown"
    for m in ["meta-llama/llama-3.3-70b-instruct:free", "deepseek/deepseek-chat-v3-0324:free"]:
        try:
            msgs = []
            if s: msgs.append({"role": "system", "content": s})
            msgs.append({"role": "user", "content": p})
            d = _http_post("https://openrouter.ai/api/v1/chat/completions",
                {"Content-Type": "application/json", "Authorization": f"Bearer {k}",
                 "HTTP-Referer": "https://pumpbom.streamlit.app"},
                {"model": m, "messages": msgs, "max_tokens": mt, "temperature": 0.1})
            r = d["choices"][0]["message"]["content"].strip()
            if r and len(r) > 10: return r
            last = "empty response"
        except Exception as e: last = str(e)
    raise Exception(f"OpenRouter failed: {last[:100]}")


def _cerebras(p, s="", mt=4000):
    k = _get_key("CEREBRAS_API_KEY")
    if not k: raise ValueError("no key")
    last = "unknown"
    for m in ["gpt-oss-120b", "llama3.1-8b"]:
        try: return _oai("https://api.cerebras.ai/v1/chat/completions", k, m, p, s, min(mt, 8000))
        except Exception as e: last = str(e)
    raise Exception(f"Cerebras failed: {last[:100]}")


# Order matters: most reliable free JSON producers first. Groq + Cerebras
# (large open models) tend to return clean JSON; Gemini is capable but flakier
# on the free tier. The chain also SKIPS any provider whose output can't be
# parsed (when want_json=True), so junk from one provider no longer blocks the
# next one (this is why "only Gemini, never Cerebras" happened before).
_PROVIDERS = [
    ("Groq", _groq), ("Cerebras", _cerebras), ("Gemini", _gemini),
    ("OpenRouter", _openrouter), ("Mistral", _mistral),
]

def _call_llm(prompt, system="", max_tokens=4000, want_json=False, extra_errors=None):
    """Try free providers in order. Returns (text, name). If want_json, a
    provider's response is only accepted when it parses as JSON — otherwise the
    chain moves to the next provider. extra_errors (e.g. a failed Claude attempt)
    are prepended to the failure detail. On total failure, raises with a
    per-provider reason so the user can see what to fix."""
    errors = list(extra_errors or [])
    for name, fn in _PROVIDERS:
        try:
            r = fn(prompt, system, max_tokens)
            if not (r and len(r.strip()) > 10):
                errors.append(f"{name}: empty response"); continue
            if want_json and _parse_json(r) is None:
                errors.append(f"{name}: unparseable (not JSON)"); continue
            return r, name
        except Exception as e:
            msg = str(e)
            if "no key" in msg.lower():
                errors.append(f"{name}: no key set")
            else:
                errors.append(f"{name}: {msg[:80]}")
    detail = " | ".join(errors) if errors else "no providers configured"
    raise Exception(
        "All free LLM providers failed/returned junk. Set/refresh an API key in "
        "Streamlit Secrets. "
        f"Details — {detail}")


def _claude_models():
    """Models to try, in order. A CLAUDE_MODEL secret (if set) takes priority.
    Current (2026) models are listed first — older dated IDs are retired and
    404 on new accounts, so they're only kept as last-resort fallbacks."""
    override = _get_key("CLAUDE_MODEL")
    base = [
        "claude-sonnet-4-6",            # current Sonnet (best value)
        "claude-haiku-4-5-20251001",    # current Haiku (cheapest)
        "claude-opus-4-8",              # current Opus (premium)
        "claude-sonnet-4-5",
        "claude-3-7-sonnet-latest",
        "claude-3-5-sonnet-latest",
        "claude-sonnet-4-20250514",     # older fallbacks (may be retired)
        "claude-3-5-haiku-20241022",
    ]
    return ([override] + base) if override else base


def _call_claude(prompt, system="", max_tokens=4000, use_search=False):
    """Claude API with optional web search. Tries multiple model IDs so an
    account that lacks one model still works on another."""
    k = _get_key("ANTHROPIC_API_KEY")
    if not k: raise ValueError("ANTHROPIC_API_KEY not set")
    try: import anthropic
    except ImportError: raise ValueError("pip install anthropic")
    client = anthropic.Anthropic(api_key=k)
    last_err = "no model tried"
    for model in _claude_models():
        kw = {"model": model, "max_tokens": min(max_tokens, 4096),
              "messages": [{"role": "user", "content": prompt}]}
        if system: kw["system"] = system
        if use_search: kw["tools"] = [{"type": "web_search_20250305", "name": "web_search"}]
        for attempt in range(3):
            try:
                resp = client.messages.create(**kw)
                u = getattr(resp, "usage", None)
                if u:
                    _track(getattr(u, "input_tokens", 0), getattr(u, "output_tokens", 0),
                           getattr(u, "cache_read_input_tokens", 0) or 0, model)
                return "\n".join(b.text for b in resp.content if hasattr(b, "text")).strip()
            except Exception as e:
                es = str(e)
                if "not_found" in es or "404" in es or "does not support" in es:
                    last_err = f"{model}: {es[:80]}"
                    break  # model unavailable on this account → try the next one
                if "429" in es or "rate_limit" in es.lower():
                    time.sleep(25 * (attempt + 1)); continue
                raise
    raise Exception(f"No available Claude model (set CLAUDE_MODEL secret). Last: {last_err}")


def _smart_call(prompt, system="", max_tokens=4000, want_json=False):
    """Claude first (if key set), else free LLMs. Returns (text, provider).
    If want_json, Claude output is only accepted when it parses as JSON.
    A failed Claude attempt's reason is surfaced in the final error."""
    pre = []
    if _get_key("ANTHROPIC_API_KEY"):
        try:
            r = _call_claude(prompt, system, max_tokens)
            if r and len(r) > 10 and (not want_json or _parse_json(r) is not None):
                return r, "Claude"
            pre.append("Claude: empty/unparseable response")
        except Exception as e:
            pre.append(f"Claude: {str(e)[:120]}")
    return _call_llm(prompt, system, max_tokens, want_json, pre)


def _cheap_call(prompt, system="", max_tokens=4000, want_json=False):
    """Cheapest path first (free LLMs), Claude only as last resort.
    Used for low-stakes steps like initial identification."""
    try:
        return _call_llm(prompt, system, max_tokens, want_json)
    except Exception:
        if _get_key("ANTHROPIC_API_KEY"):
            r = _call_claude(prompt, system, max_tokens)
            if r and len(r) > 10:
                return r, "Claude"
        raise


def _grounded_call(prompt, system="", max_tokens=2000):
    """Search-grounded call for live pricing. Preference chain:
    Claude web_search → Gemini Google-Search grounding → free LLM
    knowledge (floor-price guided). Returns (text, provider)."""
    if _get_key("ANTHROPIC_API_KEY"):
        try:
            r = _call_claude(prompt, system, max_tokens, use_search=True)
            if r and len(r) > 10: return r, "Claude+search"
        except: pass
    if _get_key("GEMINI_API_KEY"):
        try:
            r = _gemini_grounded(prompt, system, max_tokens)
            if r and len(r) > 10: return r, "Gemini+grounding"
        except: pass
    # Last resort: free LLM from knowledge (floor prices in prompt anchor it)
    return _call_llm(prompt, system, max_tokens)


# ═══════════════════════════════════════════════════════════════════
# JSON PARSER — handles fences, preamble, truncated arrays
# ═══════════════════════════════════════════════════════════════════

def _parse_json(text):
    if not text: return None
    c = text.strip()
    for f in ["```json", "```JSON", "```", "`"]: c = c.replace(f, "")
    c = c.strip()
    # Strip preamble
    for i, line in enumerate(c.split('\n')):
        if line.strip().startswith('{') or line.strip().startswith('['):
            c = '\n'.join(c.split('\n')[i:]); break
    # 1. Direct parse
    try: return json.loads(c)
    except: pass
    # 2. Try complete {...} dict
    r = _bracket_extract(c, '{', '}')
    if r is not None and isinstance(r, dict) and len(r) >= 1:
        return r
    # 3. Try complete [...] array
    r = _bracket_extract(c, '[', ']')
    if r is not None: return r
    # 4. TRUNCATED ARRAY — extract every complete {...} from broken [
    idx = c.find('[')
    if idx != -1:
        items = _recover_truncated(c[idx:])
        if items and len(items) >= 1: return items
    # 5. Single dict fallback (last resort)
    r = _bracket_extract(c, '{', '}')
    if r is not None: return r
    return None


def _bracket_extract(text, o, c):
    start = text.find(o)
    if start == -1: return None
    depth = 0; in_str = False; i = start
    while i < len(text):
        ch = text[i]
        if ch == '"' and (i == 0 or text[i-1] != '\\'): in_str = not in_str; i += 1; continue
        if in_str: i += 1; continue
        if ch == o: depth += 1
        elif ch == c:
            depth -= 1
            if depth == 0:
                s = text[start:i+1]
                try: return json.loads(s)
                except:
                    try: return json.loads(re.sub(r',\s*([}\]])', r'\1', s))
                    except: return None
        i += 1
    return None


def _recover_truncated(text):
    objects = []; i = 0
    while i < len(text):
        s = text.find('{', i)
        if s == -1: break
        depth = 0; in_str = False; j = s; found = False
        while j < len(text):
            ch = text[j]
            if ch == '"' and (j == 0 or text[j-1] != '\\'): in_str = not in_str; j += 1; continue
            if in_str: j += 1; continue
            if ch == '{': depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    try: obj = json.loads(text[s:j+1])
                    except:
                        try: obj = json.loads(re.sub(r',\s*([}\]])', r'\1', text[s:j+1]))
                        except: obj = None
                    if isinstance(obj, dict): objects.append(obj)
                    found = True; i = j + 1; break
            j += 1
        if not found: break
    return objects


# ═══════════════════════════════════════════════════════════════════
# PDF TEXT EXTRACTION
# ═══════════════════════════════════════════════════════════════════

def extract_pdf_text(file_bytes):
    """Extract a CLEAN, LLM-friendly representation of the datasheet.

    Instead of dumping flattened text (which mangles spec tables), this does a
    deterministic Python pre-pass that pulls out:
      - key:value spec pairs (lines like 'Flow : 250 m3/h', 'Battery: Li-ion')
      - tables (rendered as pipe rows)
      - the raw running text
    and packs them into one structured block. Generic — no product assumptions.
    The LLM then reads pre-digested structure, which makes IDENTIFY / SCHEMA /
    BOM far more consistent than parsing raw garble.
    """
    try:
        import pdfplumber
    except Exception as e:
        return "", str(e)
    pages, tables, kv = [], [], {}
    try:
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            for p in pdf.pages:
                t = p.extract_text()
                if t: pages.append(t)
                try:
                    for tbl in (p.extract_tables() or []):
                        rows = [[("" if c is None else str(c).strip()) for c in row]
                                for row in tbl if any(c not in (None, "") for c in row)]
                        if len(rows) >= 1 and any(len(r) >= 2 for r in rows):
                            tables.append(rows)
                except Exception:
                    pass
    except Exception as e:
        return "", str(e)

    raw_text = "\n".join(pages)
    if not raw_text.strip() and not tables:
        return "", "no extractable text"

    kv = _extract_key_values(raw_text)
    structured = _render_structured(kv, tables, raw_text)
    return structured, None


def _extract_key_values(text):
    """Pull 'Label : value' style spec pairs from datasheet text. Generic."""
    kv = {}
    for line in text.splitlines():
        line = line.strip()
        if not line or len(line) > 120:
            continue
        # split on first colon (datasheet specs) — keep short, value-bearing pairs
        if ":" in line:
            k, v = line.split(":", 1)
            k, v = k.strip(), v.strip()
            if 1 <= len(k) <= 45 and 1 <= len(v) <= 70 and not k.endswith((".", "?")) \
               and any(ch.isalnum() for ch in v) and k.lower() not in kv:
                kv[k.lower()] = f"{k}: {v}"
    # cap to avoid bloating the prompt
    return dict(list(kv.items())[:60])


def _render_structured(kv, tables, raw_text):
    parts = []
    if kv:
        parts.append("=== KEY SPECIFICATIONS (auto-extracted) ===")
        parts.extend(kv.values())
    if tables:
        parts.append("\n=== TABLES (auto-extracted) ===")
        for i, rows in enumerate(tables[:12], 1):
            parts.append(f"-- Table {i} --")
            for r in rows[:30]:
                parts.append(" | ".join(r))
    parts.append("\n=== FULL DOCUMENT TEXT ===")
    parts.append(raw_text[:16000])
    return "\n".join(parts)


# ═══════════════════════════════════════════════════════════════════
# AGENT LOGGING
# ═══════════════════════════════════════════════════════════════════

_T0 = [0.0]

def _stamp():
    el = int(time.time() - _T0[0])
    return f"[{el//60:02d}:{el%60:02d}]"


def _log(agent_log, cb, step, action, result="", running=False):
    """Append a structured entry and push a terminal-style line to UI."""
    entry = {"step": step, "action": action, "result": result, "t": _stamp()}
    agent_log.append(entry)
    mark = "◈" if running else "✓"
    line = f"{entry['t']} {mark} {step:<10} {result or action}"
    if cb:
        try: cb(line, agent_log)
        except: pass
    return entry


# ═══════════════════════════════════════════════════════════════════
# STEP 1 — IDENTIFY
# ═══════════════════════════════════════════════════════════════════

IDENTIFY_SYS = """You are a senior mechanical / procurement engineer for Indian EPC projects.
You read any engineered-product technical document (datasheet, GA drawing,
spec sheet) and identify the product. Never guess wildly — if unsure, say so
in key_specs. Return strict JSON only."""


def _identify(pdf_text, agent_log, cb):
    _log(agent_log, cb, "IDENTIFY", "Reading document structure...", running=True)
    prompt = f"""Identify the engineered product described in this document.
Read carefully — the SAME word can mean different products (e.g. "pump" may be a
large centrifugal process pump OR a tiny diaphragm sampling pump inside a handheld
gas detector). Judge from the WHOLE document (size, weight, power source, sensors,
display, wireless, housing) what the product really is.

DOCUMENT TEXT:
{pdf_text[:14000]}

Return ONLY this JSON object:
{{
  "equipment_type": "specific product class incl. standard/class if present (e.g. 'Centrifugal Pump (API 610 OH2)', 'Portable Multi-Gas Detector', 'Shell & Tube Heat Exchanger (TEMA AES)', 'Distribution Transformer')",
  "manufacturer": "name or null",
  "model": "model/tag or null",
  "is_engineered_product": true,
  "key_specs": {{ "any": "important rating/size/duty/feature parameters you find as key:value" }}
}}"""
    raw, prov = "", "none"
    try:
        raw, prov = _cheap_call(prompt, IDENTIFY_SYS, 1500, want_json=True)
    except Exception as e:
        _log(agent_log, cb, "IDENTIFY", "llm error", str(e)[:140])
    data = _parse_json(raw) or {}
    if not isinstance(data, dict): data = {}
    et = (data.get("equipment_type") or "").strip()
    # Retry with a stronger model if identification failed or was generic.
    if not et or et.lower() in ("engineered product", "unknown", "n/a", "product"):
        if raw and not et:
            _log(agent_log, cb, "IDENTIFY", "unparseable", f"got {len(raw)} chars, no equipment_type")
        try:
            raw2, prov = _smart_call(prompt, IDENTIFY_SYS, 1500, want_json=True)
            d2 = _parse_json(raw2)
            if isinstance(d2, dict) and (d2.get("equipment_type") or "").strip():
                data = d2
                et = data["equipment_type"].strip()
        except Exception as e:
            _log(agent_log, cb, "IDENTIFY", "retry failed", str(e)[:140])
    if not et:
        et = "Engineered Product"
    data["equipment_type"] = et
    data["_provider"] = prov
    if not isinstance(data.get("key_specs"), dict):
        data["key_specs"] = {}
    _log(agent_log, cb, "IDENTIFY", "done", f'equipment_type = "{et}"')
    return data


# ═══════════════════════════════════════════════════════════════════
# STEP 2 — SCHEMA (dynamic, product-specific)
# ═══════════════════════════════════════════════════════════════════

SCHEMA_SYS = """You are a polymath design engineer who can decompose ANY engineered
product in the world into its real sub-assemblies — regardless of industry or
discipline. You are equally fluent in mechanical, rotating, static, electrical,
electronic/PCB, embedded firmware, power/battery, communications, instrumentation,
optical, hydraulic, pneumatic, chemical/process, structural/civil, marine,
aerospace, automotive, medical-device and consumer-product engineering.

For the given product you FIRST reason about what kind of product it actually is
and which engineering disciplines it involves, THEN enumerate the sub-assemblies
exactly as they would appear on that product's real manufacturing/procurement BOM.

Hard rules:
- The schema MUST be specific to THIS product. A pump, a transformer, a telescope,
  a hydraulic press, a gas monitor, an EOT crane and a packaging machine have
  COMPLETELY different sub-assemblies.
- Do NOT default to mechanical/rotating groups out of habit. Derive the groups
  from the actual product.
- Include every discipline this product genuinely needs, and none that it doesn't.
- Never output a generic catch-all template.
Return strict JSON array only."""


def _build_schema(equipment_type, key_specs, agent_log, cb, doc=""):
    _log(agent_log, cb, "SCHEMA", "Building sub-assembly schema...", running=True)
    prompt = f"""Product: {equipment_type}
Known specs: {json.dumps(key_specs or {}, default=str)[:1500]}

ACTUAL DATASHEET TEXT (decompose THIS specific product, not a generic category):
{doc[:6000]}

Step 1 (think): based on the datasheet above, what kind of engineered product is
this, and which engineering disciplines does it involve? (mechanical? electronic?
hydraulic? optical? chemical/process? structural? a mix?) Beware: the same word
(e.g. "pump") can mean very different things depending on the product.

Step 2: list ALL sub-assemblies (functional groups) THIS specific product
contains, derived from what the product actually is — not from a fixed template.
Be exhaustive for this product and include every discipline it needs.

For reference, different products decompose very differently, e.g.:
- a centrifugal pump → hydraulics, rotating assembly, bearings, sealing, drive, structural
- a portable gas monitor → sensors, main PCB, power/battery, display, wireless, enclosure
- a power transformer → core, windings, tank, bushings, cooling, OLTC, protection
- a shell & tube heat exchanger → shell, tube bundle, channels, tubesheets, baffles, nozzles
- a hydraulic press → frame, cylinder, power pack, valves, controls, tooling
- a telescope/optical instrument → optics, opto-mechanics, mount/drive, electronics, housing
Use the pattern that fits THIS product; invent the right groups for it.

Return ONLY a JSON array:
[
  {{"id": "A", "name": "specific sub-assembly name",
    "description": "what it contains",
    "typical_components_count": 5}}
]
Use single-letter ids A, B, C, ... in order. 6-14 sub-assemblies typical."""
    try:
        raw, prov = _smart_call(prompt, SCHEMA_SYS, 2500, want_json=True)
    except Exception as e:
        # LLM totally unavailable — degrade to neutral schema, do not crash.
        _log(agent_log, cb, "SCHEMA", "llm unavailable", f"using fallback schema ({str(e)[:60]})")
        raw = ""
    data = _parse_json(raw)
    # Free LLMs often wrap the array in an object — unwrap it.
    if isinstance(data, dict):
        for k in ("sub_assemblies", "subassemblies", "schema", "groups", "data", "items"):
            if isinstance(data.get(k), list):
                data = data[k]; break
        else:
            # any list value, else single dict -> one-item list
            lists = [v for v in data.values() if isinstance(v, list)]
            data = lists[0] if lists else ([data] if data.get("name") else [])
    schema = []
    if isinstance(data, list):
        for i, s in enumerate(data):
            if not isinstance(s, dict): continue
            sid = str(s.get("id") or chr(65 + i)).strip() or chr(65 + i)
            name = str(s.get("name") or s.get("sub_assembly") or f"Group {sid}").strip()
            schema.append({
                "id": sid, "name": name,
                "description": str(s.get("description", "")).strip(),
                "typical_components_count": int(s.get("typical_components_count", 4) or 4),
            })
    if not schema:
        # Discipline-neutral fallback (NOT mechanical-biased) so the loop can
        # still proceed even if schema generation failed to parse.
        schema = [
            {"id": "A", "name": "Primary Functional Assembly", "description": "core parts that deliver the product's main function", "typical_components_count": 6},
            {"id": "B", "name": "Electronics & Control", "description": "PCBs, microcontroller, signal/power electronics", "typical_components_count": 5},
            {"id": "C", "name": "Power System", "description": "battery, power supply, charging", "typical_components_count": 3},
            {"id": "D", "name": "Sensing / Drive Subsystem", "description": "sensors or moving/driven parts as applicable", "typical_components_count": 4},
            {"id": "E", "name": "User Interface & Communication", "description": "display, keypad, indicators, wireless/ports", "typical_components_count": 4},
            {"id": "F", "name": "Enclosure & Structural", "description": "housing, frame, supports, sealing", "typical_components_count": 4},
            {"id": "G", "name": "Final Assembly", "description": "fasteners, gaskets, labels, finishing, assembly", "typical_components_count": 4},
        ]
    _log(agent_log, cb, "SCHEMA", "done", f"{len(schema)} sub-assemblies generated")
    return schema


# ═══════════════════════════════════════════════════════════════════
# STEP 3 — BOM GENERATION (per sub-assembly)
# ═══════════════════════════════════════════════════════════════════

BOM_SYS = """You are an expert product engineer in India producing a detailed
manufacturing/procurement BOM for ANY engineered product — mechanical, electrical,
electronic, or instrument. For each sub-assembly you list its real components.
For each component give: a specific name; its material/spec (for electronics use
the part class, e.g. 'FR4 PCB', 'Li-ion 3.7V cell', 'ARM Cortex MCU', 'graphic
LCD module'); realistic weight in kg (electronic parts are light — grams);
quantity; and applicable standards if any. type is "manufactured" (made from raw
material — castings, fabrications, machined parts, bare PCBs) or "bought_out"
(procured complete — motors, bearings, seals, sensors, MCUs, displays, batteries,
modules, connectors, fasteners). Return strict JSON array only, no prose."""


def _populate_subassembly(equipment_type, key_specs, sub, agent_log, cb, total, idx, doc=""):
    _log(agent_log, cb, "BOM",
         f"Populating: {sub['id']}. {sub['name']} ({idx}/{total})...", running=True)
    prompt = f"""Product: {equipment_type}
Specs: {json.dumps(key_specs or {}, default=str)[:1200]}

ACTUAL DATASHEET TEXT (components must fit THIS product, at THIS size/scale):
{doc[:4500]}

Sub-assembly to populate:
id={sub['id']} | name={sub['name']} | scope={sub.get('description','')}

List every real component in THIS sub-assembly, consistent with the datasheet
above. Use exact MOC grades / part classes and realistic kg weights for this
product's actual size/duty (a handheld device has gram-scale parts; a process
skid has kg/tonne-scale parts — match the real product).

Return ONLY a JSON array:
[{{"description":"component name (<=45 chars)","material":"ASTM/IS/EN grade or 'bought-out item'",
"qty":"1","unit":"no","type":"manufactured|bought_out","weight_kg":0,
"standards_applicable":"std or empty"}}]
Aim for ~{sub.get('typical_components_count',4)} components. JSON array ONLY."""
    try:
        raw, prov = _smart_call(prompt, BOM_SYS, 2500, want_json=True)
    except Exception as e:
        # Surface WHY — do not fail silently (this is what made "no components"
        # impossible to diagnose). Record the reason for the run summary.
        _populate_subassembly.last_error = str(e)
        _log(agent_log, cb, "BOM", "populate failed", f"{sub['id']}: {str(e)[:90]}")
        return []
    data = _parse_json(raw)
    if not data:
        _populate_subassembly.last_error = f"unparseable response (got {len(raw or '')} chars)"
        _log(agent_log, cb, "BOM", "populate empty", f"{sub['id']}: response not JSON")
    items = []
    if isinstance(data, list):
        items = [x for x in data if isinstance(x, dict)]
    elif isinstance(data, dict):
        for k in ["components", "bom", "items", "data"]:
            if isinstance(data.get(k), list):
                items = [x for x in data[k] if isinstance(x, dict)]; break
        else:
            if "description" in data: items = [data]
    # Recover from truncation if too few
    if len(items) < 1 and raw:
        items = _recover_truncated(raw.replace("```json", "").replace("```", ""))
    norm = []
    for c in items:
        ctype = str(c.get("type", "")).lower()
        if ctype not in ("manufactured", "bought_out"):
            ctype = _classify(c.get("description", ""))
        norm.append({
            "description": str(c.get("description", c.get("component", c.get("name", "")))).strip()[:60],
            "material": str(c.get("material", c.get("moc", ""))).strip(),
            "qty": str(c.get("qty", c.get("quantity", "1"))).strip() or "1",
            "unit": str(c.get("unit", "no")).strip() or "no",
            "type": ctype,
            "weight_kg": _num(c.get("weight_kg", c.get("weight"))),
            "standards_applicable": str(c.get("standards_applicable", c.get("standards", ""))).strip(),
            "sub_assembly_id": sub["id"],
            "sub_assembly_name": sub["name"],
        })
    return norm


def _generate_bom_batch(equipment_type, key_specs, schema, agent_log, cb, doc=""):
    """Generate the WHOLE BOM in ONE LLM call (all sub-assemblies at once).
    ~7-10× fewer calls than per-sub-assembly → big token/cost saving.
    Returns a normalized list, or [] if the single call didn't produce enough."""
    _log(agent_log, cb, "BOM", "Generating full BOM (single call)...", running=True)
    sub_list = [{"id": s["id"], "name": s["name"], "scope": s.get("description", ""),
                 "approx_count": s.get("typical_components_count", 4)} for s in schema]
    prompt = f"""Product: {equipment_type}
Specs: {json.dumps(key_specs or {}, default=str)[:1000]}

ACTUAL DATASHEET TEXT:
{doc[:7000]}

Sub-assemblies to populate (cover EVERY one):
{json.dumps(sub_list, default=str)}

Produce the COMPLETE BOM in a SINGLE JSON array. For each sub-assembly include its
real components (use sub_assembly_id to tag each). Exact MOC grades / part classes
and realistic kg weights for THIS product's actual size/scale.

Return ONLY a JSON array (no prose):
[{{"sub_assembly_id":"A","description":"component (<=45 chars)","material":"grade or 'bought-out item'",
"qty":"1","unit":"no","type":"manufactured|bought_out","weight_kg":0,"standards_applicable":""}}]"""
    try:
        raw, prov = _smart_call(prompt, BOM_SYS, 4096, want_json=True)
    except Exception as e:
        _populate_subassembly.last_error = str(e)
        _log(agent_log, cb, "BOM", "single-call failed", str(e)[:90])
        return []
    data = _parse_json(raw)
    if isinstance(data, dict):
        for k in ("components", "bom", "items", "data"):
            if isinstance(data.get(k), list): data = data[k]; break
        else: data = []
    if not isinstance(data, list):
        data = _recover_truncated(raw.replace("```json", "").replace("```", "")) if raw else []
    sub_by_id = {s["id"]: s for s in schema}
    norm = []
    for c in data:
        if not isinstance(c, dict) or not c.get("description"): continue
        sid = str(c.get("sub_assembly_id", "")).strip()
        sub = sub_by_id.get(sid) or schema[0]
        ctype = str(c.get("type", "")).lower()
        if ctype not in ("manufactured", "bought_out"):
            ctype = _classify(c.get("description", ""))
        norm.append({
            "description": str(c.get("description", "")).strip()[:60],
            "material": str(c.get("material", c.get("moc", ""))).strip(),
            "qty": str(c.get("qty", "1")).strip() or "1",
            "unit": str(c.get("unit", "no")).strip() or "no",
            "type": ctype,
            "weight_kg": _num(c.get("weight_kg", c.get("weight"))),
            "standards_applicable": str(c.get("standards_applicable", c.get("standards", ""))).strip(),
            "sub_assembly_id": sub["id"], "sub_assembly_name": sub["name"],
        })
    return norm


def _generate_bom(equipment_type, key_specs, schema, agent_log, cb, doc=""):
    _populate_subassembly.last_error = ""
    n = len(schema)
    # 1) Try the cheap single-call path first.
    bom = _generate_bom_batch(equipment_type, key_specs, schema, agent_log, cb, doc)
    covered = len({c["sub_assembly_id"] for c in bom})
    # 2) Fall back to per-sub-assembly only if the batch was thin/incomplete.
    if len(bom) < max(5, n) or covered < max(1, n // 2):
        _log(agent_log, cb, "BOM", "expanding per sub-assembly...", running=True)
        bom = []
        for i, sub in enumerate(schema, 1):
            bom.extend(_populate_subassembly(equipment_type, key_specs, sub,
                                             agent_log, cb, n, i, doc))
    for j, c in enumerate(bom, 1):
        c["id"] = j
    if not bom:
        reason = getattr(_populate_subassembly, "last_error", "") or "LLM returned no usable components"
        _log(agent_log, cb, "BOM", "NO COMPONENTS", f"reason: {reason[:110]}")
    else:
        _log(agent_log, cb, "BOM", "done", f"{len(bom)} components across {n} sub-assemblies")
    return bom


# ═══════════════════════════════════════════════════════════════════
# STEP 3.5 — COMPLETENESS CRITIC (agentic self-review)
# A second pass: re-read the datasheet + the BOM we built, and ask what
# STANDARD/mandatory components are missing for THIS product. Generalist —
# no per-product hardcoding; the model applies the relevant standard itself.
# ═══════════════════════════════════════════════════════════════════

VERIFY_SYS = """You are a strict QA / checking engineer reviewing a Bill of Materials
for completeness against applicable engineering standards and normal build practice
for the specific product. You know the mandatory components every well-formed BOM of
that product type must contain (e.g. API 610 pump → casing, impeller, wear rings,
shaft, sleeve, bearings, seal, coupling, baseplate; a portable electronic instrument
→ main PCB, MCU, battery, display, sensors, enclosure, etc.). You ONLY report what is
genuinely MISSING — do not duplicate components already present. Return strict JSON."""


def _verify_completeness(equipment_type, key_specs, schema, bom, agent_log, cb, doc=""):
    """Critic pass: find standard/mandatory components missing from the BOM and add
    them. Returns (bom, n_added)."""
    _log(agent_log, cb, "VERIFY", "Reviewing BOM against standards...", running=True)
    present = sorted({str(c.get("description", "")).strip() for c in bom if c.get("description")})
    sub_list = [{"id": s["id"], "name": s["name"]} for s in schema]
    prompt = f"""Product: {equipment_type}
Specs: {json.dumps(key_specs or {}, default=str)[:800]}

DATASHEET (ground truth):
{doc[:4000]}

Sub-assemblies (use these ids for placement):
{json.dumps(sub_list, default=str)}

The CURRENT BOM already contains these components:
{json.dumps(present, default=str)[:3500]}

Task: per applicable standards and normal build practice for THIS product, list ONLY
the MANDATORY components that are MISSING from the list above. Do not repeat anything
already present. If nothing is missing, return [].

Return ONLY a JSON array:
[{{"description":"component (<=45 chars)","material":"grade or 'bought-out item'",
"qty":"1","unit":"no","type":"manufactured|bought_out","weight_kg":0,
"standards_applicable":"std or empty","sub_assembly_id":"best-fit id from list"}}]"""
    try:
        raw, prov = _smart_call(prompt, VERIFY_SYS, 2500, want_json=True)
    except Exception as e:
        _log(agent_log, cb, "VERIFY", "skipped", f"critic unavailable ({str(e)[:50]})")
        return bom, 0
    data = _parse_json(raw)
    if isinstance(data, dict):
        for k in ("missing", "components", "items", "data"):
            if isinstance(data.get(k), list):
                data = data[k]; break
        else:
            data = [data] if data.get("description") else []
    if not isinstance(data, list):
        data = []

    present_lc = {p.lower() for p in present}
    sub_by_id = {s["id"]: s for s in schema}
    added = 0
    for c in data:
        if not isinstance(c, dict):
            continue
        desc = str(c.get("description", "")).strip()[:60]
        if not desc or desc.lower() in present_lc:
            continue
        sid = str(c.get("sub_assembly_id", "")).strip()
        sub = sub_by_id.get(sid) or (schema[-1] if schema else {"id": "Z", "name": "Additional Items"})
        ctype = str(c.get("type", "")).lower()
        if ctype not in ("manufactured", "bought_out"):
            ctype = _classify(desc)
        bom.append({
            "description": desc,
            "material": str(c.get("material", c.get("moc", ""))).strip(),
            "qty": str(c.get("qty", "1")).strip() or "1",
            "unit": str(c.get("unit", "no")).strip() or "no",
            "type": ctype,
            "weight_kg": _num(c.get("weight_kg", c.get("weight"))),
            "standards_applicable": str(c.get("standards_applicable", c.get("standards", ""))).strip(),
            "sub_assembly_id": sub["id"],
            "sub_assembly_name": sub["name"],
            "added_by": "completeness_critic",
        })
        present_lc.add(desc.lower())
        added += 1
    for j, c in enumerate(bom, 1):
        c["id"] = j
    _log(agent_log, cb, "VERIFY", "done",
         f"{added} missing component(s) added" if added else "no gaps found")
    return bom, added


# ═══════════════════════════════════════════════════════════════════
# STEP 4 — VALIDATION LOOP (agentic)
# ═══════════════════════════════════════════════════════════════════

def _validate_bom(bom, equipment_type, schema):
    """Returns {completeness, gaps, warnings}."""
    gaps, warnings = [], []
    by_sub = {}
    for c in bom:
        by_sub.setdefault(c.get("sub_assembly_id"), []).append(c)

    populated = 0
    for sub in schema:
        cnt = len(by_sub.get(sub["id"], []))
        if cnt == 0:
            gaps.append(f"{sub['id']}. {sub['name']} — no components")
        else:
            populated += 1
            if cnt < max(2, int(sub.get("typical_components_count", 4) * 0.5)):
                warnings.append(f"{sub['id']}. {sub['name']} — only {cnt} components (low)")

    schema_cov = populated / max(len(schema), 1)

    # Reasonableness of total count
    total = len(bom)
    if total < 10:
        warnings.append(f"Total components ({total}) low for {equipment_type}")
    size_ok = 1.0 if total >= 15 else (total / 15.0)

    completeness = round(0.7 * schema_cov + 0.3 * size_ok, 3)
    return {"completeness": completeness, "gaps": gaps, "warnings": warnings,
            "populated_subs": populated, "total_components": total}


def _fill_gaps(equipment_type, key_specs, schema, bom, gaps, agent_log, cb, doc=""):
    """Re-populate sub-assemblies flagged as empty/low."""
    gap_ids = set()
    for g in gaps:
        gid = g.split(".")[0].strip()
        if gid: gap_ids.add(gid)
    n = len(gap_ids)
    if not n: return bom
    sub_by_id = {s["id"]: s for s in schema}
    for k, gid in enumerate(sorted(gap_ids), 1):
        sub = sub_by_id.get(gid)
        if not sub: continue
        comps = _populate_subassembly(equipment_type, key_specs, sub, agent_log, cb, n, k, doc)
        # remove any existing items for this sub then re-add
        bom = [c for c in bom if c.get("sub_assembly_id") != gid]
        bom.extend(comps)
    for j, c in enumerate(bom, 1):
        c["id"] = j
    return bom


# ═══════════════════════════════════════════════════════════════════
# STEP 5 — SHOULD-COST PRICING
# RM (gross weight × live ₹/kg) + machining. No overhead, no margin.
# ═══════════════════════════════════════════════════════════════════

COST_SYS = """You are a cost engineer at an Indian EPC company doing should-cost analysis
for ANY engineered-product component. Find ACTUAL CURRENT manufacturing cost —
not selling price, not textbook price.

Use live market data where possible. Search:
- SAIL/RINL price lists for steel grades
- LME India for non-ferrous metals
- IndiaMART / TradeIndia for castings, machined parts, bought-out items

Use REALISTIC 2025-26 Indian prices. Reference bands (₹/kg, finished):
- Grey cast iron casting: 180-220
- IS 2062 fabricated steel: 130-160
- SS316 / CF8M casting: 750-950
- High chrome iron A532: 900-1200
- Alloy steel bar (EN19/EN24, machined): 400-600
- CNC machining shop rate: 1200-1800/hr
- LT 415V motor: 5000-8000/kW ; HT motor: 4000-5500/kW

All amounts INR. No overhead. No margin. Floor manufacturing cost only."""


def _classify(comp, moc=""):
    c = (comp or "").upper()
    if "MOTOR" in c and not any(x in c for x in ["BOLT","PULLEY","SIDE","BRACKET","HOUSING","MOUNT"]): return "bought_out"
    if "BEARING" in c and "HOUSING" not in c: return "bought_out"
    if any(x in c for x in ["MECHANICAL SEAL","MECH SEAL","SEAL KIT"]): return "bought_out"
    if any(x in c for x in ["V-BELT","V BELT","VBELT"]): return "bought_out"
    if "COMPANION FLANGE" in c or "COUNTER FLANGE" in c: return "bought_out"
    if "GASKET" in c or "O-RING" in c or "O RING" in c: return "bought_out"
    if "FOUNDATION" in c and "BOLT" in c: return "bought_out"
    if any(x in c for x in ["RTD","THERMOMETER","PRESSURE GAUGE","TRANSMITTER","SWITCH","SENSOR","PT100","GAUGE"]): return "bought_out"
    if any(x in c for x in ["FIRST FILL","GREASE","LUBRICANT","OIL FILL"]): return "bought_out"
    if "GUARD" in c: return "bought_out"
    if any(x in c for x in ["NUT","BOLT","STUD","WASHER","FASTENER"]): return "bought_out"
    return "manufactured"


def _num(v):
    try:
        if v in (None, "", "null", "None"): return 0.0
        return float(re.findall(r"-?\d+\.?\d*", str(v))[0])
    except: return 0.0


def _price_bom(equipment_type, key_specs, bom, agent_log, cb):
    """Add should-cost columns to each component dict in place. Returns bom."""
    if not bom: return bom
    _log(agent_log, cb, "PRICE", "Starting should-cost...", running=True)

    items = []
    for c in bom:
        ctype = c.get("type") or _classify(c.get("description", ""))
        items.append({"c": c, "cat": ctype})

    mfg = [x for x in items if x["cat"] == "manufactured"]
    bo = [x for x in items if x["cat"] == "bought_out"]

    spec_str = json.dumps(key_specs or {}, default=str)[:600]

    # ---- Manufactured: RM (gross weight × ₹/kg) + machining ----
    for i in range(0, len(mfg), 3):
        batch = mfg[i:i+3]
        _log(agent_log, cb, "PRICE",
             "fetching market rates: " + ", ".join(x["c"]["description"][:18] for x in batch),
             running=True)
        items_j = json.dumps([{"id": x["c"]["id"], "description": x["c"]["description"],
                               "material": x["c"]["material"], "weight_kg": x["c"]["weight_kg"],
                               "qty": x["c"]["qty"]} for x in batch], default=str)
        prompt = f"""Should-cost for these {equipment_type} components. Specs: {spec_str}

Components:
{items_j}

For EACH component:
1. Find live ₹/kg for its exact material grade (India 2025-26).
2. Gross weight = net weight × 1.35 for castings, × 1.15 for forgings/bar/plate.
3. raw_material_cost = gross_weight × ₹/kg.
4. machining_cost: realistic — CNC turning/boring ₹1200-1600/hr, pattern+mould ₹20-25/kg gross, fabrication welding ₹130-160/kg.
5. total = raw + machining. No overhead, no margin.

Floor prices if a rate is unavailable: grey CI ₹190/kg, SS316 ₹850/kg, CS/WCB ₹220/kg, alloy steel bar ₹480/kg, high chrome ₹1000/kg, MS fabricated ₹145/kg.

Return JSON array ONLY:
[{{"id":<n>,"raw_material_rate_per_kg":<int>,"gross_weight_kg":<num>,
"raw_material_cost_inr":<int>,"machining_cost_inr":<int>,"total_cost_inr":<int>,
"material_source":"site/basis","confidence":"high|medium|low","notes":"calc basis"}}]"""
        try:
            raw, prov = _grounded_call(prompt, COST_SYS, 2000)
            parsed = _parse_json(raw)
            if not isinstance(parsed, list): parsed = [parsed] if isinstance(parsed, dict) else []
            pm = {p["id"]: p for p in parsed if isinstance(p, dict) and "id" in p}
            for x in batch:
                p = pm.get(x["c"]["id"], {})
                raw_c = int(_num(p.get("raw_material_cost_inr")))
                mach = int(_num(p.get("machining_cost_inr")))
                total = int(_num(p.get("total_cost_inr")) or (raw_c + mach))
                _apply_price(x["c"], raw_c, mach, total, "manufactured",
                             p.get("confidence", "medium"), p.get("material_source", prov),
                             f"₹{p.get('raw_material_rate_per_kg',0)}/kg × {p.get('gross_weight_kg',0)}kg | {p.get('notes','')}")
        except Exception as e:
            for x in batch:
                _apply_price(x["c"], 0, 0, 0, "manufactured", "error", str(e)[:80], "")
        if i + 3 < len(mfg): time.sleep(_pricing_delay())

    # ---- Bought-out: market procurement price ----
    for i in range(0, len(bo), 4):
        batch = bo[i:i+4]
        _log(agent_log, cb, "PRICE",
             "market pricing: " + ", ".join(x["c"]["description"][:18] for x in batch),
             running=True)
        items_j = json.dumps([{"id": x["c"]["id"], "description": x["c"]["description"],
                               "material": x["c"]["material"], "qty": x["c"]["qty"]}
                              for x in batch], default=str)
        prompt = f"""Find OEM MARKET PROCUREMENT PRICE (what the fabricator pays) for these
bought-out items of a {equipment_type}. Specs: {spec_str}

Items:
{items_j}

Use the right search per item (motor, bearing, seal, gasket, instrument, fastener,
flange, gearbox, coupling, etc.). Give realistic current Indian price each.

Return JSON array ONLY:
[{{"id":<n>,"market_price_inr":<int>,"source":"IndiaMART/TradeIndia/OEM","confidence":"high|medium|low","notes":"make/size referenced"}}]"""
        try:
            raw, prov = _grounded_call(prompt, COST_SYS, 2000)
            parsed = _parse_json(raw)
            if not isinstance(parsed, list): parsed = [parsed] if isinstance(parsed, dict) else []
            pm = {p["id"]: p for p in parsed if isinstance(p, dict) and "id" in p}
            for x in batch:
                p = pm.get(x["c"]["id"], {})
                price = int(_num(p.get("market_price_inr")))
                _apply_price(x["c"], price, 0, price, "bought_out",
                             p.get("confidence", "medium"), p.get("source", prov),
                             p.get("notes", ""))
        except Exception as e:
            for x in batch:
                _apply_price(x["c"], 0, 0, 0, "bought_out", "error", str(e)[:80], "")
        if i + 4 < len(bo): time.sleep(_pricing_delay())

    _log(agent_log, cb, "PRICE", "done", "should-cost complete")
    return bom


def _pricing_delay():
    # Claude TPM limits need long gaps; free grounding can go faster.
    if _get_key("ANTHROPIC_API_KEY"): return 35
    if _get_key("GEMINI_API_KEY"): return 8
    return 2


def _apply_price(c, raw_c, mach, total, ctype, conf, source, notes):
    qty = max(_num(c.get("qty")) or 1, 1)
    line_total = int(total * qty)
    c.update({
        "raw_material_inr": raw_c,
        "machining_inr": mach,
        "unit_cost_inr": total,
        "total_cost_inr": line_total,
        "gst_18pct": int(line_total * 0.18),
        "price_with_gst": int(line_total * 1.18),
        "price_confidence": conf,
        "price_source": str(source)[:120],
        "price_notes": str(notes)[:200],
        "component_type": ctype,
    })


# ═══════════════════════════════════════════════════════════════════
# MANUAL / USER-DRIVEN PRICING (no LLM, no web search)
# Buyer enters raw-material ₹/kg per component; labour via slider;
# supplier (Indian/International) scales the cost. Deterministic.
# ═══════════════════════════════════════════════════════════════════

SUPPLIER_FACTORS = {"Indian": 1.0, "International": 1.4}


def _gross_factor(c):
    """Net→gross weight multiplier inferred from the part's nature."""
    blob = (str(c.get("description", "")) + " " + str(c.get("material", ""))).lower()
    if any(k in blob for k in ["cast", "casing", "volute", "impeller", "housing"]):
        return 1.35
    if any(k in blob for k in ["forg", " bar", "shaft", "machined", "sleeve", "gear"]):
        return 1.15
    return 1.10


def price_manual(bom, rate_map, labour_rate_per_kg=60.0, supplier="Indian", gst=0.18):
    """Deterministic costing from buyer inputs.

    rate_map: {component_id: raw_material_₹_per_kg}
    labour_rate_per_kg: slider value (₹/kg of gross weight)
    supplier: 'Indian' or 'International' (applies SUPPLIER_FACTORS)

    cost = (raw_material + labour) × supplier_factor, per unit × qty.
    Returns (bom, should_cost_summary).
    """
    factor = SUPPLIER_FACTORS.get(supplier, 1.0)
    for c in bom:
        net = _num(c.get("weight_kg"))
        qty = max(_num(c.get("qty")) or 1, 1)
        gross = round(net * _gross_factor(c), 3)
        rate = _num(rate_map.get(str(c.get("id")), rate_map.get(c.get("id"), 0)))
        raw = gross * rate
        labour = gross * _num(labour_rate_per_kg)
        unit = (raw + labour) * factor
        total = unit * qty
        c.update({
            "gross_weight_kg": gross,
            "raw_material_rate": int(rate),
            "raw_material_inr": int(raw * factor),
            "machining_inr": int(labour * factor),
            "unit_cost_inr": int(unit),
            "total_cost_inr": int(total),
            "gst_18pct": int(total * gst),
            "price_with_gst": int(total * (1 + gst)),
            "price_confidence": "user",
            "price_source": f"{supplier} | ₹{int(rate)}/kg RM + ₹{int(_num(labour_rate_per_kg))}/kg labour ×{factor}",
            "price_notes": f"gross {gross}kg",
            "component_type": c.get("type", c.get("component_type", "")),
            "supplier_type": supplier,
        })
    return bom, _build_should_cost(bom)


def _build_should_cost(bom):
    if not bom: return {}
    def s(k): return int(sum(_num(c.get(k)) for c in bom))
    sub_totals = {}
    for c in bom:
        key = f"{c.get('sub_assembly_id','?')}. {c.get('sub_assembly_name','')}"
        sub_totals[key] = sub_totals.get(key, 0) + int(_num(c.get("total_cost_inr")))
    sub_totals = dict(sorted(sub_totals.items(), key=lambda kv: kv[1], reverse=True))
    top = sorted(bom, key=lambda c: _num(c.get("total_cost_inr")), reverse=True)[:5]
    top5 = [{"description": c.get("description"), "total_cost_inr": int(_num(c.get("total_cost_inr"))),
             "sub_assembly": c.get("sub_assembly_name"), "confidence": c.get("price_confidence")}
            for c in top]
    conf = {}
    for c in bom:
        k = c.get("price_confidence", "n/a"); conf[k] = conf.get(k, 0) + 1
    types = {}
    for c in bom:
        k = c.get("component_type", "n/a"); types[k] = types.get(k, 0) + 1
    return {
        "total_raw_material": s("raw_material_inr"),
        "total_machining": s("machining_inr"),
        "total_ex_gst": s("total_cost_inr"),
        "total_gst": s("gst_18pct"),
        "total_incl_gst": s("price_with_gst"),
        "sub_totals": {k: int(v) for k, v in sub_totals.items()},
        "top5_drivers": top5,
        "confidence": conf,
        "type_split": types,
        "component_count": len(bom),
        "note": "Should-Cost = Raw Material (live rates) + Machining. NO overhead, NO margin. PO − should-cost = supplier profit.",
    }


# ═══════════════════════════════════════════════════════════════════
# STEP 6 — CONFIDENCE
# ═══════════════════════════════════════════════════════════════════

def _confidence(bom, schema, validation):
    if not bom: return 0.0
    schema_completeness = validation.get("completeness", 0.0)
    priced = [c for c in bom if _num(c.get("total_cost_inr")) > 0]
    pricing_coverage = len(priced) / max(len(bom), 1)
    def has_mat(c):
        m = (c.get("material") or "").strip().lower()
        return bool(m) and m not in ("null", "none", "n/a", "bought-out item", "-")
    material_spec_quality = sum(1 for c in bom if has_mat(c)) / max(len(bom), 1)
    standards_coverage = sum(1 for c in bom if (c.get("standards_applicable") or "").strip()) / max(len(bom), 1)
    score = (0.40 * schema_completeness + 0.30 * pricing_coverage +
             0.20 * material_spec_quality + 0.10 * standards_coverage)
    return round(score, 3)


# ═══════════════════════════════════════════════════════════════════
# AGENT LOOP — orchestrates all steps
# ═══════════════════════════════════════════════════════════════════

def run_agent(pdf_text, progress_callback=None, price=True):
    """
    Full agentic loop. Claude/free LLMs decide schema and components; the
    engine validates and loops to fill gaps, then prices and scores.

    Returns: {
        "equipment_type": str, "manufacturer": str, "model": str,
        "schema": [{"id","name","description","typical_components_count"}],
        "bom": [component dicts with pricing],
        "should_cost": {...},
        "confidence": float,
        "agent_log": [{"step","action","result","t"}],
        "gaps": [...], "warnings": [...],
        "iterations": int,
    }
    """
    _T0[0] = time.time()
    _reset_usage()
    agent_log, cb = [], progress_callback
    doc = pdf_text or ""

    def _toks():
        s = usage_snapshot()
        _log(agent_log, cb, "TOKENS", "usage",
             f"{s['calls']} calls · {s['total_tokens']:,} tokens · ~₹{s['est_cost_inr']}")

    # 1 — IDENTIFY
    ident = _identify(pdf_text, agent_log, cb)
    equipment_type = ident.get("equipment_type", "Engineered Product")
    key_specs = ident.get("key_specs", {}) or {}

    # 2 — SCHEMA (grounded in the actual datasheet text)
    schema = _build_schema(equipment_type, key_specs, agent_log, cb, doc)

    # 3 — BOM (grounded in the actual datasheet text)
    bom = _generate_bom(equipment_type, key_specs, schema, agent_log, cb, doc)

    # 3.5 — COMPLETENESS CRITIC (self-review against standards)
    try:
        bom, _added = _verify_completeness(equipment_type, key_specs, schema, bom, agent_log, cb, doc)
    except Exception as e:
        _log(agent_log, cb, "VERIFY", "skipped", str(e)[:60])

    # 4 — VALIDATION LOOP (one pass; refill only genuinely EMPTY sub-assemblies
    #     to keep cost down — re-populating low-count subs is rarely worth the tokens)
    iterations = 0
    validation = _validate_bom(bom, equipment_type, schema)
    if validation["completeness"] < 0.80 and validation["gaps"]:
        iterations = 1
        _log(agent_log, cb, "VALIDATE",
             f"completeness {int(validation['completeness']*100)}% → filling {len(validation['gaps'])} empty sub-assemblies",
             f"{int(validation['completeness']*100)}% — filling {len(validation['gaps'])} gaps")
        bom = _fill_gaps(equipment_type, key_specs, schema, bom,
                         validation["gaps"], agent_log, cb, doc)
        validation = _validate_bom(bom, equipment_type, schema)
    _log(agent_log, cb, "VALIDATE", "done",
         f"completeness {int(validation['completeness']*100)}% after {iterations} iteration(s)")

    # 5 — PRICE (optional; new blueprint prices manually in the UI)
    if not price:
        _log(agent_log, cb, "PRICE", "skipped", "manual costing in UI (buyer enters rates)")
    try:
        if price:
            bom = _price_bom(equipment_type, key_specs, bom, agent_log, cb)
    except Exception as e:
        _log(agent_log, cb, "PRICE", "error", str(e)[:120])
        for c in bom:
            if "total_cost_inr" not in c:
                _apply_price(c, 0, 0, 0, c.get("type", "manufactured"), "error", "pricing failed", "")
    should_cost = _build_should_cost(bom)

    # 6 — CONFIDENCE
    confidence = _confidence(bom, schema, validation)
    usage = usage_snapshot()
    _log(agent_log, cb, "TOKENS", "total",
         f"{usage['calls']} calls · {usage['total_tokens']:,} tokens "
         f"(in {usage['input']:,} / out {usage['output']:,}) · ~₹{usage['est_cost_inr']}")
    _log(agent_log, cb, "DONE", "done",
         f"confidence {int(confidence*100)}% | {len(bom)} components | ₹{should_cost.get('total_ex_gst',0):,}")

    return {
        "equipment_type": equipment_type,
        "manufacturer": ident.get("manufacturer"),
        "model": ident.get("model"),
        "key_specs": key_specs,
        "schema": schema,
        "bom": bom,
        "should_cost": should_cost,
        "confidence": confidence,
        "usage": usage,
        "agent_log": agent_log,
        "gaps": validation["gaps"],
        "warnings": validation["warnings"],
        "iterations": iterations,
    }


# ═══════════════════════════════════════════════════════════════════
# DATAFRAME HELPERS
# ═══════════════════════════════════════════════════════════════════

def bom_to_dataframe(bom, priced=None):
    """Convert agent bom list to a display DataFrame. Auto-detects pricing."""
    if not bom: return pd.DataFrame()
    has_price = priced if priced is not None else any("total_cost_inr" in c for c in bom)
    rows = []
    for c in bom:
        if not isinstance(c, dict): continue
        row = {
            "No": c.get("id", ""),
            "Sub_Assembly": f"{c.get('sub_assembly_id','')}. {c.get('sub_assembly_name','')}".strip(". "),
            "Component": c.get("description", ""),
            "Material": c.get("material", ""),
            "Qty": c.get("qty", "1"),
            "Unit": c.get("unit", "no"),
            "Weight_kg": c.get("weight_kg", 0),
            "Type": c.get("type", c.get("component_type", "")),
            "Standards": c.get("standards_applicable", ""),
        }
        if has_price:
            row.update({
                "Raw_Material_INR": int(_num(c.get("raw_material_inr"))),
                "Machining_INR": int(_num(c.get("machining_inr"))),
                "Total_Price_INR": int(_num(c.get("total_cost_inr"))),
                "Price_Confidence": c.get("price_confidence", ""),
                "Price_Source": c.get("price_source", ""),
                "Price_Notes": c.get("price_notes", ""),
            })
        rows.append(row)
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════
# EXCEL EXPORT — dynamic sub-assemblies + Agent Summary sheet
# ═══════════════════════════════════════════════════════════════════

def export_excel(result):
    """result = output of run_agent. Returns BytesIO xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime

    bom = result.get("bom", [])
    sc = result.get("should_cost", {})
    df = bom_to_dataframe(bom)
    priced = "Total_Price_INR" in df.columns

    wb = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: Agent Summary ──────────────────────────────────────
    ws0 = wb.active; ws0.title = "Agent Summary"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 30; ws0.column_dimensions["B"].width = 58
    ws0.merge_cells("A1:B1")
    c = ws0["A1"]; c.value = "AGENTIC BOM — SHOULD-COST SUMMARY"
    c.font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="0a0a0f")
    c.alignment = Alignment(horizontal="center", vertical="center")
    conf = result.get("confidence", 0)
    summary = [
        ("Equipment Type", result.get("equipment_type", "")),
        ("Manufacturer", result.get("manufacturer") or "—"),
        ("Model", result.get("model") or "—"),
        ("Confidence Score", f"{int(conf*100)}%"),
        ("Total Components", str(sc.get("component_count", len(bom)))),
        ("Sub-assemblies", str(len(result.get("schema", [])))),
        ("Agent Iterations", str(result.get("iterations", 0))),
        ("Gaps Flagged", str(len(result.get("gaps", [])))),
        ("Total Raw Material (₹)", f"{sc.get('total_raw_material',0):,}"),
        ("Total Machining (₹)", f"{sc.get('total_machining',0):,}"),
        ("Should-Cost ex-GST (₹)", f"{sc.get('total_ex_gst',0):,}"),
        ("GST 18% (₹)", f"{sc.get('total_gst',0):,}"),
        ("Should-Cost incl-GST (₹)", f"{sc.get('total_incl_gst',0):,}"),
        ("Date", datetime.date.today().isoformat()),
    ]
    gt = result.get("grand_total")
    if gt:
        summary += [
            ("Components (RM+Mfg) (₹)", f"{gt.get('components_ex_gst',0):,}"),
            ("Freight (₹)", f"{gt.get('freight',0):,}"),
            ("Overhead (₹)", f"{gt.get('overhead',0):,}"),
            ("GRAND TOTAL ex-GST (₹)", f"{gt.get('total_ex_gst',0):,}"),
            ("GRAND TOTAL incl-GST (₹)", f"{gt.get('total_incl_gst',0):,}"),
        ]
    r = 3
    for lbl, val in summary:
        ws0.cell(r, 1, lbl).font = Font(name="Arial", bold=True, size=10)
        ws0.cell(r, 1).fill = PatternFill("solid", fgColor="EEF2F7"); ws0.cell(r, 1).border = bdr
        ws0.cell(r, 2, val).font = Font(name="Arial", size=10); ws0.cell(r, 2).border = bdr
        r += 1
    if result.get("gaps"):
        r += 1
        ws0.cell(r, 1, "Gaps / Warnings").font = Font(name="Arial", bold=True, size=10, color="C0392B")
        r += 1
        for g in result.get("gaps", []) + result.get("warnings", []):
            ws0.cell(r, 1, "•").font = Font(name="Arial", size=9)
            ws0.cell(r, 2, str(g)).font = Font(name="Arial", size=9); r += 1
    ws0.cell(r + 1, 1, sc.get("note", "")).font = Font(name="Arial", italic=True, size=8, color="666666")

    # ── Sheet 2: BOM (grouped by dynamic sub-assembly) ──────────────
    ws1 = wb.create_sheet("BOM"); ws1.sheet_view.showGridLines = False
    if priced:
        cols = ["No", "Sub_Assembly", "Component", "Material", "Qty", "Unit", "Weight_kg",
                "Type", "Raw_Material_INR", "Machining_INR", "Total_Price_INR",
                "Price_Confidence", "Price_Notes"]
    else:
        cols = ["No", "Sub_Assembly", "Component", "Material", "Qty", "Unit",
                "Weight_kg", "Type", "Standards"]
    cols = [x for x in cols if x in df.columns]
    widths = {"No": 5, "Sub_Assembly": 26, "Component": 30, "Material": 22, "Qty": 6,
              "Unit": 6, "Weight_kg": 10, "Type": 13, "Standards": 20, "Raw_Material_INR": 13,
              "Machining_INR": 13, "Total_Price_INR": 14, "Price_Confidence": 11, "Price_Notes": 30}

    ncols = max(len(cols), 1)   # guard: openpyxl columns are 1-indexed (0 is invalid)
    ws1.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws1["A1"]; t.value = f"BILL OF MATERIALS — {result.get('equipment_type','')}"
    t.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="0a0a0f")

    if df.empty or not cols:
        # No components generated — write a clear note instead of crashing.
        ws1.cell(3, 1, "No components were generated for this run "
                       "(BOM empty — check the Agent Log sheet for why).").font = \
            Font(name="Arial", italic=True, size=10, color="C0392B")
        ws1.column_dimensions["A"].width = 70
    else:
        rr = 2
        for j, col in enumerate(cols):
            h = ws1.cell(rr, j + 1, col.replace("_", " "))
            h.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            h.fill = PatternFill("solid", fgColor="4a7a9b")
            h.alignment = Alignment(horizontal="center", wrap_text=True); h.border = bdr
            ws1.column_dimensions[get_column_letter(j + 1)].width = widths.get(col, 14)
        rr += 1

        # group rows by sub-assembly, in schema order
        order = {f"{s['id']}. {s['name']}": i for i, s in enumerate(result.get("schema", []))}
        df2 = df.copy()
        df2["_ord"] = df2["Sub_Assembly"].map(lambda s: order.get(s, 99))
        df2 = df2.sort_values(["_ord", "No"]).drop(columns=["_ord"])

        f1, f2 = PatternFill("solid", fgColor="EEF4FB"), PatternFill("solid", fgColor="FFFFFF")
        grp = PatternFill("solid", fgColor="DDE6EF")
        last_sub = None
        i = 0
        for _, row in df2.iterrows():
            sub = row.get("Sub_Assembly", "")
            if sub != last_sub:
                ws1.merge_cells(f"A{rr}:{get_column_letter(ncols)}{rr}")
                g = ws1.cell(rr, 1, sub)
                g.font = Font(name="Arial", bold=True, size=9, color="1F2A36")
                g.fill = grp; g.border = bdr
                rr += 1; last_sub = sub
            for j, col in enumerate(cols):
                v = row.get(col, "")
                if pd.isna(v): v = ""
                if col in ("Raw_Material_INR", "Machining_INR", "Total_Price_INR"):
                    try: v = f"₹{int(float(v)):,}" if v != "" else ""
                    except: pass
                cell = ws1.cell(rr, j + 1, v)
                cell.font = Font(name="Arial", size=8); cell.fill = f1 if i % 2 == 0 else f2
                cell.border = bdr; cell.alignment = Alignment(wrap_text=True, vertical="top")
            rr += 1; i += 1
        ws1.freeze_panes = "A3"


    # ── Sheet 3: Should-Cost ────────────────────────────────────────
    ws2 = wb.create_sheet("Should-Cost"); ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 40; ws2.column_dimensions["B"].width = 20
    ws2.merge_cells("A1:B1")
    t2 = ws2["A1"]; t2.value = "SHOULD-COST BREAKDOWN"
    t2.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t2.fill = PatternFill("solid", fgColor="3d6b4f")
    rr = 3
    ws2.cell(rr, 1, "Sub-assembly").font = Font(bold=True, size=10)
    ws2.cell(rr, 2, "Cost (₹ ex-GST)").font = Font(bold=True, size=10)
    rr += 1
    for k, v in sc.get("sub_totals", {}).items():
        ws2.cell(rr, 1, k).font = Font(size=9); ws2.cell(rr, 1).border = bdr
        ws2.cell(rr, 2, f"₹{int(v):,}").font = Font(size=9); ws2.cell(rr, 2).border = bdr
        rr += 1
    rr += 1
    for lbl, key in [("Total Raw Material", "total_raw_material"),
                     ("Total Machining", "total_machining"),
                     ("Total ex-GST", "total_ex_gst"),
                     ("GST 18%", "total_gst"),
                     ("Total incl-GST", "total_incl_gst")]:
        ws2.cell(rr, 1, lbl).font = Font(bold=True, size=10)
        ws2.cell(rr, 2, f"₹{int(sc.get(key,0)):,}").font = Font(bold=True, size=10)
        rr += 1

    # ── Sheet 4: Agent Log ──────────────────────────────────────────
    ws3 = wb.create_sheet("Agent Log"); ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 10; ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 70
    for j, h in enumerate(["Time", "Step", "Detail"], 1):
        cell = ws3.cell(1, j, h); cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="e8a020")
    rr = 2
    for e in result.get("agent_log", []):
        ws3.cell(rr, 1, e.get("t", "")).font = Font(name="Consolas", size=9)
        ws3.cell(rr, 2, e.get("step", "")).font = Font(name="Consolas", size=9)
        ws3.cell(rr, 3, e.get("result") or e.get("action", "")).font = Font(name="Consolas", size=9)
        rr += 1

    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf
